#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
使用 pdfplumber 提取银行对账单 PDF 表格并转换为 Excel
"""

import pdfplumber
import openpyxl
import os
import re
from pathlib import Path

# PDF 文件目录
pdf_dir = Path("/home/admin/.openclaw/media/inbound")
# 输出 Excel 文件
output_excel = "/home/admin/openclaw/workspace/银行对账单汇总.xlsx"

# 获取所有对账单 PDF 文件
pdf_files = sorted([f for f in pdf_dir.glob("BOCVC442981877088*.pdf")])

print(f"找到 {len(pdf_files)} 个 PDF 文件")

# 创建 Excel 工作簿
wb = openpyxl.Workbook()
wb.remove(wb.active)

all_transactions = []

for pdf_idx, pdf_path in enumerate(pdf_files):
    print(f"\n[{pdf_idx + 1}/{len(pdf_files)}] 处理：{pdf_path.name}")
    
    parts = pdf_path.stem.split("-")
    if len(parts) >= 3:
        year = parts[1]
        seq = parts[2].lstrip("0")
        sheet_name = f"{year}年{seq}月"
    else:
        sheet_name = pdf_path.stem[:15]
    sheet_name = sheet_name[:31]
    
    ws = wb.create_sheet(title=sheet_name)
    
    with pdfplumber.open(pdf_path) as pdf:
        print(f"  共 {len(pdf.pages)} 页")
        
        row_num = 1
        headers_set = False
        
        for page_num, page in enumerate(pdf.pages):
            tables = page.extract_tables()
            print(f"  第{page_num+1}页：{len(tables)} 个表格")
            
            for table in tables:
                for row in table:
                    if row is None:
                        continue
                    
                    cleaned = [str(c).strip() if c else "" for c in row]
                    
                    # 检查表头
                    if not headers_set and any("序号" in c for c in cleaned):
                        headers = ["序号", "记账日", "起息日", "交易类型", "凭证", "用途/摘要", 
                                   "借方发生额", "贷方发生额", "余额", "机构/柜员", "备注"]
                        for ci, h in enumerate(headers, 1):
                            ws.cell(row=row_num, column=ci, value=h)
                        row_num += 1
                        headers_set = True
                        continue
                    
                    if headers_set and len(cleaned) >= 5:
                        for ci, v in enumerate(cleaned, 1):
                            if v:
                                ws.cell(row=row_num, column=ci, value=v)
                        
                        trans = {
                            "文件": pdf_path.name,
                            "期间": sheet_name,
                            "序号": cleaned[0] if len(cleaned) > 0 else "",
                            "记账日": cleaned[1] if len(cleaned) > 1 else "",
                            "起息日": cleaned[2] if len(cleaned) > 2 else "",
                            "交易类型": cleaned[3] if len(cleaned) > 3 else "",
                            "凭证": cleaned[4] if len(cleaned) > 4 else "",
                            "用途摘要": cleaned[5] if len(cleaned) > 5 else "",
                            "借方发生额": cleaned[6] if len(cleaned) > 6 else "",
                            "贷方发生额": cleaned[7] if len(cleaned) > 7 else "",
                            "余额": cleaned[8] if len(cleaned) > 8 else "",
                            "机构柜员": cleaned[9] if len(cleaned) > 9 else "",
                            "备注": cleaned[10] if len(cleaned) > 10 else "",
                        }
                        all_transactions.append(trans)
                        row_num += 1
        
        print(f"  写入 {row_num-1} 行")

# 创建汇总表
summary_ws = wb.create_sheet(title="全部交易汇总", index=0)
headers = ["文件", "期间", "序号", "记账日", "起息日", "交易类型", "凭证", "用途摘要", 
           "借方发生额", "贷方发生额", "余额", "机构柜员", "备注"]
for ci, h in enumerate(headers, 1):
    summary_ws.cell(row=1, column=ci, value=h)

for ri, t in enumerate(all_transactions, 2):
    summary_ws.cell(row=ri, column=1, value=t["文件"])
    summary_ws.cell(row=ri, column=2, value=t["期间"])
    summary_ws.cell(row=ri, column=3, value=t["序号"])
    summary_ws.cell(row=ri, column=4, value=t["记账日"])
    summary_ws.cell(row=ri, column=5, value=t["起息日"])
    summary_ws.cell(row=ri, column=6, value=t["交易类型"])
    summary_ws.cell(row=ri, column=7, value=t["凭证"])
    summary_ws.cell(row=ri, column=8, value=t["用途摘要"])
    summary_ws.cell(row=ri, column=9, value=t["借方发生额"])
    summary_ws.cell(row=ri, column=10, value=t["贷方发生额"])
    summary_ws.cell(row=ri, column=11, value=t["余额"])
    summary_ws.cell(row=ri, column=12, value=t["机构柜员"])
    summary_ws.cell(row=ri, column=13, value=t["备注"])

wb.save(output_excel)
print(f"\n✅ 完成！共 {len(all_transactions)} 条记录")
print(f"   文件：{output_excel}")
print(f"   大小：{os.path.getsize(output_excel)/1024:.1f} KB")
