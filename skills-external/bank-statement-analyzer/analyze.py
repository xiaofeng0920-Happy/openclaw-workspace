#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
银行对账单分析技能 - 核心处理脚本
"""

import subprocess
import openpyxl
import re
import os
import sys
from pathlib import Path
from datetime import datetime
from collections import defaultdict

class BankStatementAnalyzer:
    """银行对账单分析器"""
    
    def __init__(self, pdf_dir, output_dir):
        self.pdf_dir = Path(pdf_dir)
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)
        self.all_data = []
        self.pdf_files = []
        
    def find_pdfs(self, start_date=None, end_date=None):
        """查找对账单 PDF 文件"""
        pattern = 'BOC*.pdf' if 'BOC' in str(self.pdf_dir) else '*.pdf'
        all_pdfs = sorted(self.pdf_dir.glob(pattern))
        
        if start_date or end_date:
            filtered = []
            for p in all_pdfs:
                # 从文件名提取日期
                parts = p.stem.split('-')
                if len(parts) >= 3:
                    year = parts[1]
                    month = parts[2].lstrip('0')
                    # 简单过滤逻辑
                    filtered.append(p)
            self.pdf_files = filtered
        else:
            self.pdf_files = all_pdfs
            
        print(f'找到 {len(self.pdf_files)} 个 PDF 文件')
        return self.pdf_files
    
    def extract_pdf_text(self, pdf_path):
        """提取 PDF 文本"""
        try:
            # 尝试 pdftotext
            result = subprocess.run(
                ['pdftotext', '-layout', str(pdf_path), '-'],
                capture_output=True, text=True, timeout=30
            )
            return result.stdout
        except:
            # 备用方案：pdfplumber
            try:
                import pdfplumber
                with pdfplumber.open(pdf_path) as pdf:
                    text = ''
                    for page in pdf.pages:
                        text += page.extract_text() or ''
                    return text
            except Exception as e:
                print(f'提取失败 {pdf_path.name}: {e}')
                return ''
    
    def parse_text(self, text, filename, period):
        """解析文本提取交易记录"""
        records = []
        
        for line in text.split('\n'):
            if '|' not in line or '序号' in line or '合计' in line:
                continue
            
            # 匹配数据行
            match = re.search(r'\|\s*(\d+)\s*\|', line)
            if match:
                parts = [x.strip() for x in line.split('|')]
                if len(parts) >= 5:
                    record = {
                        '文件': filename,
                        '期间': period,
                        '序号': parts[1] if len(parts) > 1 else '',
                        '记账日': parts[2] if len(parts) > 2 else '',
                        '起息日': parts[3] if len(parts) > 3 else '',
                        '交易类型': parts[4] if len(parts) > 4 else '',
                        '凭证': parts[5] if len(parts) > 5 else '',
                        '用途摘要': parts[6] if len(parts) > 6 else '',
                        '借方发生额': parts[7] if len(parts) > 7 else '',
                        '贷方发生额': parts[8] if len(parts) > 8 else '',
                        '余额': parts[9] if len(parts) > 9 else '',
                        '机构': parts[10] if len(parts) > 10 else '',
                        '备注': parts[11] if len(parts) > 11 else '',
                    }
                    records.append(record)
        
        return records
    
    def process_all(self):
        """处理所有 PDF"""
        for i, pdf in enumerate(self.pdf_files, 1):
            print(f'[{i}/{len(self.pdf_files)}] {pdf.name}')
            
            # 提取期间信息
            parts = pdf.stem.split('-')
            if len(parts) >= 3:
                year = parts[1]
                month = parts[2].lstrip('0')
                period = f'{year}年{month}月'
            else:
                period = 'unknown'
            
            # 提取文本
            text = self.extract_pdf_text(pdf)
            if not text:
                continue
            
            # 解析数据
            records = self.parse_text(text, pdf.name, period)
            self.all_data.extend(records)
            print(f'  提取 {len(records)} 条记录')
        
        print(f'\n共提取 {len(self.all_data)} 条记录')
        return self.all_data
    
    def save_excel(self, filename='银行对账单汇总.xlsx'):
        """保存 Excel 文件"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # 按月创建工作表
        monthly_data = defaultdict(list)
        for record in self.all_data:
            monthly_data[record['期间']].append(record)
        
        for period, records in sorted(monthly_data.items()):
            ws = wb.create_sheet(period[:31])
            
            # 表头
            headers = ['序号', '记账日', '起息日', '交易类型', '凭证', 
                      '用途摘要', '借方发生额', '贷方发生额', '余额', '机构', '备注']
            for c, h in enumerate(headers, 1):
                ws.cell(1, c, h)
            
            # 数据
            for r, rec in enumerate(records, 2):
                ws.cell(r, 1, rec['序号'])
                ws.cell(r, 2, rec['记账日'])
                ws.cell(r, 3, rec['起息日'])
                ws.cell(r, 4, rec['交易类型'])
                ws.cell(r, 5, rec['凭证'])
                ws.cell(r, 6, rec['用途摘要'])
                ws.cell(r, 7, rec['借方发生额'])
                ws.cell(r, 8, rec['贷方发生额'])
                ws.cell(r, 9, rec['余额'])
                ws.cell(r, 10, rec['机构'])
                ws.cell(r, 11, rec['备注'])
        
        # 汇总表
        sws = wb.create_sheet('汇总', index=0)
        h2 = ['文件', '期间', '序号', '记账日', '起息日', '交易类型', '凭证',
              '用途摘要', '借方发生额', '贷方发生额', '余额', '机构', '备注']
        for c, h in enumerate(h2, 1):
            sws.cell(1, c, h)
        
        for r, rec in enumerate(self.all_data, 2):
            sws.cell(r, 1, rec['文件'])
            sws.cell(r, 2, rec['期间'])
            sws.cell(r, 3, rec['序号'])
            sws.cell(r, 4, rec['记账日'])
            sws.cell(r, 5, rec['起息日'])
            sws.cell(r, 6, rec['交易类型'])
            sws.cell(r, 7, rec['凭证'])
            sws.cell(r, 8, rec['用途摘要'])
            sws.cell(r, 9, rec['借方发生额'])
            sws.cell(r, 10, rec['贷方发生额'])
            sws.cell(r, 11, rec['余额'])
            sws.cell(r, 12, rec['机构'])
            sws.cell(r, 13, rec['备注'])
        
        output_path = self.output_dir / filename
        wb.save(output_path)
        size = os.path.getsize(output_path) / 1024
        print(f'✅ Excel: {filename} ({size:.0f}KB)')
        return output_path
    
    def analyze(self):
        """财务分析"""
        # 分类统计
        income_categories = defaultdict(float)
        expense_categories = defaultdict(float)
        monthly_balance = {}
        
        for rec in self.all_data:
            try:
                debit = float(str(rec['借方发生额']).replace(',', '') or '0')
                credit = float(str(rec['贷方发生额']).replace(',', '') or '0')
                summary = rec['用途摘要']
                period = rec['期间']
                
                # 月度余额
                if rec['余额']:
                    try:
                        monthly_balance[period] = float(str(rec['余额']).replace(',', ''))
                    except:
                        pass
                
                # 收入分类
                if credit > 0:
                    if any(x in summary for x in ['赎回', '分红']):
                        income_categories['基金赎回'] += credit
                    elif any(x in summary for x in ['管理费']):
                        income_categories['管理费'] += credit
                    elif any(x in summary for x in ['业绩', '提成']):
                        income_categories['业绩报酬'] += credit
                    elif any(x in summary for x in ['退税', '返还']):
                        income_categories['税费返还'] += credit
                    else:
                        income_categories['其他收入'] += credit
                
                # 支出分类
                if debit > 0:
                    if any(x in summary for x in ['申购', '投资', '认购']):
                        expense_categories['基金投资'] += debit
                    elif any(x in summary for x in ['工资', '代发']):
                        expense_categories['工资'] += debit
                    elif '税' in summary:
                        expense_categories['税费'] += debit
                    elif any(x in summary for x in ['房租', '租金']):
                        expense_categories['房租'] += debit
                    elif '物业' in summary:
                        expense_categories['物业费'] += debit
                    elif '报销' in summary:
                        if '兰鹏' in summary:
                            expense_categories['兰鹏报销'] += debit
                        else:
                            expense_categories['员工报销'] += debit
                    elif '公积金' in summary:
                        expense_categories['公积金'] += debit
                    else:
                        expense_categories['其他支出'] += debit
            except:
                pass
        
        return {
            'income': dict(income_categories),
            'expense': dict(expense_categories),
            'balance': monthly_balance,
            'total_records': len(self.all_data)
        }
    
    def generate_report(self, analysis, filename='财务分析报告.md'):
        """生成 Markdown 报告"""
        report = []
        report.append('# 银行对账单财务分析报告\n')
        report.append(f'**生成时间：** {datetime.now().strftime("%Y-%m-%d %H:%M")}\n')
        report.append(f'**分析期间：** {len(analysis["balance"])} 个月\n')
        report.append(f'**总记录数：** {analysis["total_records"]} 条\n')
        
        # 余额趋势
        report.append('\n## 📊 余额趋势\n')
        for period, balance in sorted(analysis['balance'].items()):
            report.append(f'- {period}: {balance:,.2f} 元')
        
        # 收入分析
        report.append('\n\n## 💰 收入分类\n')
        for cat, amt in sorted(analysis['income'].items(), key=lambda x: x[1], reverse=True):
            report.append(f'- {cat}: {amt:,.2f} 元')
        
        # 支出分析
        report.append('\n\n## 💸 支出分类\n')
        for cat, amt in sorted(analysis['expense'].items(), key=lambda x: x[1], reverse=True):
            report.append(f'- {cat}: {amt:,.2f} 元')
        
        # 风险提示
        report.append('\n\n## ⚠️ 风险提示\n')
        if analysis['expense'].get('兰鹏报销', 0) > 100000:
            report.append('- 兰鹏报销金额较大，建议核实\n')
        
        report.append('\n---\n*Generated by bank-statement-analyzer*')
        
        content = '\n'.join(report)
        output_path = self.output_dir / filename
        with open(output_path, 'w') as f:
            f.write(content)
        
        size = os.path.getsize(output_path) / 1024
        print(f'✅ 报告：{filename} ({size:.0f}KB)')
        return output_path


def main():
    """主函数"""
    pdf_dir = sys.argv[1] if len(sys.argv) > 1 else '/home/admin/.openclaw/media/inbound'
    output_dir = sys.argv[2] if len(sys.argv) > 2 else '/home/admin/openclaw/workspace'
    
    print('='*60)
    print('🏦 银行对账单分析技能')
    print('='*60)
    
    analyzer = BankStatementAnalyzer(pdf_dir, output_dir)
    analyzer.find_pdfs()
    analyzer.process_all()
    analyzer.save_excel()
    
    analysis = analyzer.analyze()
    analyzer.generate_report(analysis)
    
    print('\n✅ 分析完成！')


if __name__ == '__main__':
    main()
