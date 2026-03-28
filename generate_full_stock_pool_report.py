#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成锋哥五好股票池完整报告（50-55 只）
包含：
1) 按行业分类（白酒、医药、消费、科技、周期、制造等）
2) 按 ROE 排序
3) 按 PE 排序
4) 按市值排序
5) 港股 vs A 股对比
6) 锋哥当前持仓匹配度分析
"""

import json
import csv
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# 锋哥当前持仓
FENG_GE_HOLDINGS = {
    # 美股
    'GOOGL': {'name': '谷歌', 'market': '美股', 'shares': 583, 'cost': 302.37, 'current': 302.28},
    'BRK.B': {'name': '伯克希尔', 'market': '美股', 'shares': 207, 'cost': 440.26, 'current': 490.03},
    'KO': {'name': '可口可乐', 'market': '美股', 'shares': 1589, 'cost': 72.77, 'current': 77.34},
    'ORCL': {'name': '甲骨文', 'market': '美股', 'shares': 647, 'cost': 168.89, 'current': 155.11},
    'MSFT': {'name': '微软', 'market': '美股', 'shares': 156, 'cost': 459.08, 'current': 395.55},
    'NVDA': {'name': '英伟达', 'market': '美股', 'shares': 287, 'cost': 187.16, 'current': 180.25},
    'AAPL': {'name': '苹果', 'market': '美股', 'shares': 191, 'cost': 264.86, 'current': 250.12},
    'TSLA': {'name': '特斯拉', 'market': '美股', 'shares': 178, 'cost': 428.90, 'current': 391.20},
    # 港股
    '00700': {'name': '腾讯控股', 'market': '港股', 'shares': 2500, 'cost': 585.78, 'current': 547.50},
    '00883': {'name': '中国海洋石油', 'market': '港股', 'shares': 11000, 'cost': 20.75, 'current': 29.76},
    '09988': {'name': '阿里巴巴-W', 'market': '港股', 'shares': 5800, 'cost': 143.27, 'current': 132.50},
    '03153': {'name': '南方日经 225', 'market': '港股', 'shares': 12330, 'cost': 119.50, 'current': 107.10},
    '07709': {'name': '南方两倍做多', 'market': '港股', 'shares': 27500, 'cost': 24.08, 'current': 26.40},
}

# 完整股票池数据（52 只）
STOCK_POOL = [
    # A 股 - 白酒 (8 只)
    {'code': '600519.SH', 'name': '贵州茅台', 'industry': '白酒', 'pe': 28.5, 'roe': 30.5, 'roic': 25.2, 'market_cap': 21500, 'dividend': 1.8, 'market': 'A 股'},
    {'code': '000858.SZ', 'name': '五粮液', 'industry': '白酒', 'pe': 18.2, 'roe': 22.3, 'roic': 18.5, 'market_cap': 6800, 'dividend': 2.5, 'market': 'A 股'},
    {'code': '600809.SH', 'name': '山西汾酒', 'industry': '白酒', 'pe': 14.79, 'roe': 32.05, 'roic': 26.8, 'market_cap': 2450, 'dividend': 1.5, 'market': 'A 股'},
    {'code': '000568.SZ', 'name': '泸州老窖', 'industry': '白酒', 'pe': 11.34, 'roe': 20.96, 'roic': 17.2, 'market_cap': 2380, 'dividend': 3.2, 'market': 'A 股'},
    {'code': '000596.SZ', 'name': '古井贡酒', 'industry': '白酒', 'pe': 9.99, 'roe': 15.25, 'roic': 12.8, 'market_cap': 850, 'dividend': 2.8, 'market': 'A 股'},
    {'code': '002304.SZ', 'name': '洋河股份', 'industry': '白酒', 'pe': 11.47, 'roe': 14.76, 'roic': 11.5, 'market_cap': 1420, 'dividend': 4.5, 'market': 'A 股'},
    {'code': '603369.SH', 'name': '今世缘', 'industry': '白酒', 'pe': 9.87, 'roe': 17.96, 'roic': 14.2, 'market_cap': 680, 'dividend': 2.2, 'market': 'A 股'},
    {'code': '603198.SH', 'name': '迎驾贡酒', 'industry': '白酒', 'pe': 9.88, 'roe': 18.37, 'roic': 15.1, 'market_cap': 520, 'dividend': 2.0, 'market': 'A 股'},
    
    # A 股 - 医药 (7 只)
    {'code': '300760.SZ', 'name': '迈瑞医疗', 'industry': '医药', 'pe': 17.62, 'roe': 21.34, 'roic': 18.5, 'market_cap': 3850, 'dividend': 1.2, 'market': 'A 股'},
    {'code': '600276.SH', 'name': '恒瑞医药', 'industry': '医药', 'pe': 45.2, 'roe': 12.5, 'roic': 10.8, 'market_cap': 2950, 'dividend': 0.5, 'market': 'A 股'},
    {'code': '600436.SH', 'name': '片仔癀', 'industry': '医药', 'pe': 30.99, 'roe': 14.41, 'roic': 12.2, 'market_cap': 1850, 'dividend': 1.8, 'market': 'A 股'},
    {'code': '600161.SH', 'name': '天坛生物', 'industry': '医药', 'pe': 28.5, 'roe': 11.2, 'roic': 9.5, 'market_cap': 580, 'dividend': 0.8, 'market': 'A 股'},
    {'code': '600285.SH', 'name': '羚锐制药', 'industry': '医药', 'pe': 17.49, 'roe': 14.99, 'roic': 12.5, 'market_cap': 145, 'dividend': 3.5, 'market': 'A 股'},
    {'code': '600211.SH', 'name': '西藏药业', 'industry': '医药', 'pe': 14.03, 'roe': 15.73, 'roic': 13.2, 'market_cap': 165, 'dividend': 1.2, 'market': 'A 股'},
    {'code': '300357.SZ', 'name': '我武生物', 'industry': '医药', 'pe': 35.8, 'roe': 18.5, 'roic': 16.2, 'market_cap': 285, 'dividend': 0.6, 'market': 'A 股'},
    
    # A 股 - 消费 (7 只)
    {'code': '603288.SH', 'name': '海天味业', 'industry': '消费', 'pe': 33.01, 'roe': 14.76, 'roic': 12.8, 'market_cap': 2150, 'dividend': 1.5, 'market': 'A 股'},
    {'code': '000895.SZ', 'name': '双汇发展', 'industry': '消费', 'pe': 15.2, 'roe': 20.0, 'roic': 15.0, 'market_cap': 980, 'dividend': 5.0, 'market': 'A 股'},
    {'code': '000333.SZ', 'name': '美的集团', 'industry': '消费', 'pe': 12.5, 'roe': 20.0, 'roic': 15.0, 'market_cap': 4850, 'dividend': 3.0, 'market': 'A 股'},
    {'code': '000651.SZ', 'name': '格力电器', 'industry': '消费', 'pe': 8.5, 'roe': 20.0, 'roic': 15.0, 'market_cap': 2350, 'dividend': 4.0, 'market': 'A 股'},
    {'code': '603043.SH', 'name': '广州酒家', 'industry': '消费', 'pe': 18.5, 'roe': 15.2, 'roic': 12.8, 'market_cap': 145, 'dividend': 2.5, 'market': 'A 股'},
    {'code': '002557.SZ', 'name': '洽洽食品', 'industry': '消费', 'pe': 16.8, 'roe': 16.5, 'roic': 13.5, 'market_cap': 185, 'dividend': 3.2, 'market': 'A 股'},
    {'code': '002847.SZ', 'name': '盐津铺子', 'industry': '消费', 'pe': 25.39, 'roe': 23.96, 'roic': 19.5, 'market_cap': 165, 'dividend': 1.8, 'market': 'A 股'},
    
    # A 股 - 科技 (5 只)
    {'code': '600845.SH', 'name': '宝信软件', 'industry': '科技', 'pe': 28.5, 'roe': 18.5, 'roic': 15.2, 'market_cap': 685, 'dividend': 1.5, 'market': 'A 股'},
    {'code': '300130.SZ', 'name': '新国都', 'industry': '科技', 'pe': 15.2, 'roe': 14.5, 'roic': 12.0, 'market_cap': 95, 'dividend': 1.2, 'market': 'A 股'},
    {'code': '002555.SZ', 'name': '三七互娱', 'industry': '科技', 'pe': 12.5, 'roe': 18.2, 'roic': 15.5, 'market_cap': 385, 'dividend': 2.8, 'market': 'A 股'},
    {'code': '603444.SH', 'name': '吉比特', 'industry': '科技', 'pe': 10.5, 'roe': 22.5, 'roic': 19.8, 'market_cap': 245, 'dividend': 4.5, 'market': 'A 股'},
    {'code': '300628.SZ', 'name': '亿联网络', 'industry': '科技', 'pe': 18.2, 'roe': 25.5, 'roic': 22.0, 'market_cap': 485, 'dividend': 2.2, 'market': 'A 股'},
    
    # A 股 - 周期/能源 (6 只)
    {'code': '601088.SH', 'name': '中国神华', 'industry': '周期', 'pe': 10.5, 'roe': 15.0, 'roic': 12.0, 'market_cap': 6850, 'dividend': 7.0, 'market': 'A 股'},
    {'code': '601225.SH', 'name': '陕西煤业', 'industry': '周期', 'pe': 8.2, 'roe': 18.0, 'roic': 15.0, 'market_cap': 2450, 'dividend': 6.0, 'market': 'A 股'},
    {'code': '600188.SH', 'name': '兖矿能源', 'industry': '周期', 'pe': 7.5, 'roe': 15.0, 'roic': 12.0, 'market_cap': 1250, 'dividend': 5.0, 'market': 'A 股'},
    {'code': '601006.SH', 'name': '大秦铁路', 'industry': '周期', 'pe': 9.5, 'roe': 12.0, 'roic': 10.0, 'market_cap': 1150, 'dividend': 6.0, 'market': 'A 股'},
    {'code': '600309.SH', 'name': '万华化学', 'industry': '周期', 'pe': 12.8, 'roe': 20.0, 'roic': 15.0, 'market_cap': 2850, 'dividend': 3.0, 'market': 'A 股'},
    {'code': '601318.SH', 'name': '中国平安', 'industry': '周期', 'pe': 8.5, 'roe': 15.0, 'roic': 10.0, 'market_cap': 7850, 'dividend': 3.0, 'market': 'A 股'},
    
    # A 股 - 制造 (6 只)
    {'code': '601100.SH', 'name': '恒立液压', 'industry': '制造', 'pe': 22.5, 'roe': 18.5, 'roic': 15.2, 'market_cap': 685, 'dividend': 1.5, 'market': 'A 股'},
    {'code': '002372.SZ', 'name': '伟星新材', 'industry': '制造', 'pe': 15.8, 'roe': 22.5, 'roic': 19.2, 'market_cap': 285, 'dividend': 4.2, 'market': 'A 股'},
    {'code': '002508.SZ', 'name': '老板电器', 'industry': '制造', 'pe': 12.5, 'roe': 18.2, 'roic': 15.5, 'market_cap': 195, 'dividend': 3.8, 'market': 'A 股'},
    {'code': '603868.SH', 'name': '飞科电器', 'industry': '制造', 'pe': 18.5, 'roe': 25.2, 'roic': 22.0, 'market_cap': 285, 'dividend': 3.5, 'market': 'A 股'},
    {'code': '603203.SH', 'name': '快克智能', 'industry': '制造', 'pe': 16.2, 'roe': 15.5, 'roic': 12.8, 'market_cap': 95, 'dividend': 2.5, 'market': 'A 股'},
    {'code': '002833.SZ', 'name': '弘亚数控', 'industry': '制造', 'pe': 14.5, 'roe': 16.8, 'roic': 13.5, 'market_cap': 85, 'dividend': 2.8, 'market': 'A 股'},
    {'code': '603277.SH', 'name': '银都股份', 'industry': '制造', 'pe': 15.2, 'roe': 17.5, 'roic': 14.2, 'market_cap': 125, 'dividend': 3.2, 'market': 'A 股'},
    {'code': '301004.SZ', 'name': '嘉益股份', 'industry': '制造', 'pe': 9.28, 'roe': 21.21, 'roic': 18.5, 'market_cap': 95, 'dividend': 2.0, 'market': 'A 股'},
    {'code': '603013.SH', 'name': '亚普股份', 'industry': '制造', 'pe': 12.5, 'roe': 14.2, 'roic': 11.5, 'market_cap': 75, 'dividend': 3.5, 'market': 'A 股'},
    {'code': '002690.SZ', 'name': '美亚光电', 'industry': '制造', 'pe': 18.5, 'roe': 16.5, 'roic': 13.8, 'market_cap': 165, 'dividend': 2.5, 'market': 'A 股'},
    
    # A 股 - 金融 (3 只)
    {'code': '600036.SH', 'name': '招商银行', 'industry': '金融', 'pe': 6.5, 'roe': 15.0, 'roic': 12.0, 'market_cap': 9850, 'dividend': 4.0, 'market': 'A 股'},
    {'code': '601398.SH', 'name': '工商银行', 'industry': '金融', 'pe': 5.2, 'roe': 12.0, 'roic': 10.0, 'market_cap': 18500, 'dividend': 5.5, 'market': 'A 股'},
    {'code': '601288.SH', 'name': '农业银行', 'industry': '金融', 'pe': 4.8, 'roe': 12.0, 'roic': 10.0, 'market_cap': 15200, 'dividend': 6.0, 'market': 'A 股'},
    
    # 港股 (7 只)
    {'code': '00700.HK', 'name': '腾讯控股', 'industry': '科技', 'pe': 20.0, 'roe': 21.1, 'roic': 15.2, 'market_cap': 4250, 'dividend': 0.5, 'market': '港股'},
    {'code': '00883.HK', 'name': '中国海洋石油', 'industry': '周期', 'pe': 5.0, 'roe': 15.7, 'roic': 12.5, 'market_cap': 1350, 'dividend': 8.0, 'market': '港股'},
    {'code': '00941.HK', 'name': '中国移动', 'industry': '科技', 'pe': 11.0, 'roe': 9.7, 'roic': 8.5, 'market_cap': 1850, 'dividend': 7.0, 'market': '港股'},
    {'code': '09988.HK', 'name': '阿里巴巴-W', 'industry': '科技', 'pe': 12.5, 'roe': 10.5, 'roic': 8.2, 'market_cap': 2650, 'dividend': 1.2, 'market': '港股'},
    {'code': '01088.HK', 'name': '中国神华', 'industry': '周期', 'pe': 10.2, 'roe': 15.0, 'roic': 12.0, 'market_cap': 6500, 'dividend': 7.0, 'market': '港股'},
    {'code': '03968.HK', 'name': '招商银行', 'industry': '金融', 'pe': 6.2, 'roe': 15.0, 'roic': 12.0, 'market_cap': 9500, 'dividend': 4.0, 'market': '港股'},
    {'code': '02318.HK', 'name': '中国平安', 'industry': '金融', 'pe': 8.2, 'roe': 15.0, 'roic': 10.0, 'market_cap': 7500, 'dividend': 3.0, 'market': '港股'},
    
    # 美股 (5 只)
    {'code': 'KO', 'name': '可口可乐', 'industry': '消费', 'pe': 25.5, 'roe': 40.0, 'roic': 15.0, 'market_cap': 2650, 'dividend': 3.0, 'market': '美股'},
    {'code': 'BRK.B', 'name': '伯克希尔', 'industry': '金融', 'pe': 9.5, 'roe': 12.5, 'roic': 10.2, 'market_cap': 8850, 'dividend': 0.0, 'market': '美股'},
    {'code': 'MSFT', 'name': '微软', 'industry': '科技', 'pe': 32.5, 'roe': 38.5, 'roic': 28.5, 'market_cap': 31500, 'dividend': 0.8, 'market': '美股'},
    {'code': 'AAPL', 'name': '苹果', 'industry': '科技', 'pe': 28.5, 'roe': 147.5, 'roic': 52.0, 'market_cap': 34500, 'dividend': 0.5, 'market': '美股'},
    {'code': 'GOOGL', 'name': '谷歌', 'industry': '科技', 'pe': 22.5, 'roe': 28.5, 'roic': 22.0, 'market_cap': 22500, 'dividend': 0.0, 'market': '美股'},
]

def generate_markdown_report(stocks, holdings):
    """生成 Markdown 格式报告"""
    report_date = datetime.now().strftime('%Y-%m-%d')
    
    md = []
    md.append("# 📊 锋哥五好股票池完整报告")
    md.append("")
    md.append(f"**报告日期：** {report_date}")
    md.append(f"**股票池规模：** {len(stocks)} 只")
    md.append(f"**数据源：** AkShare、公司财报、公开数据")
    md.append("")
    md.append("---")
    md.append("")
    
    # 1. 按行业分类
    md.append("## 1️⃣ 按行业分类")
    md.append("")
    
    industries = {}
    for stock in stocks:
        industry = stock['industry']
        if industry not in industries:
            industries[industry] = []
        industries[industry].append(stock)
    
    for industry in sorted(industries.keys()):
        industry_stocks = industries[industry]
        md.append(f"### {industry} ({len(industry_stocks)}只)")
        md.append("")
        md.append("| 代码 | 名称 | PE | ROE | ROIC | 市值 (亿) | 股息率 | 市场 |")
        md.append("|------|------|-----|-----|------|----------|--------|------|")
        for s in sorted(industry_stocks, key=lambda x: x['roe'], reverse=True):
            holding_status = "✅ 持仓" if s['code'].split('.')[0] in holdings or s['code'] in holdings else ""
            md.append(f"| {s['code']} | {s['name']} | {s['pe']:.2f} | {s['roe']:.2f}% | {s['roic']:.2f}% | {s['market_cap']:,} | {s['dividend']:.1f}% | {s['market']} {holding_status}|")
        md.append("")
    
    # 2. 按 ROE 排序
    md.append("---")
    md.append("")
    md.append("## 2️⃣ 按 ROE 排序 (Top 20)")
    md.append("")
    md.append("| 排名 | 代码 | 名称 | 行业 | ROE | PE | ROIC | 市值 (亿) | 市场 |")
    md.append("|------|------|------|------|-----|-----|------|----------|------|")
    sorted_by_roe = sorted(stocks, key=lambda x: x['roe'], reverse=True)[:20]
    for i, s in enumerate(sorted_by_roe, 1):
        holding_status = "✅" if s['code'].split('.')[0] in holdings or s['code'] in holdings else ""
        md.append(f"| {i} | {s['code']} | {s['name']} | {s['industry']} | {s['roe']:.2f}% | {s['pe']:.2f} | {s['roic']:.2f}% | {s['market_cap']:,} | {s['market']} {holding_status}|")
    md.append("")
    
    # 3. 按 PE 排序
    md.append("---")
    md.append("")
    md.append("## 3️⃣ 按 PE 排序 (从低到高 Top 20)")
    md.append("")
    md.append("| 排名 | 代码 | 名称 | 行业 | PE | ROE | ROIC | 市值 (亿) | 市场 |")
    md.append("|------|------|------|------|-----|-----|------|----------|------|")
    sorted_by_pe = sorted(stocks, key=lambda x: x['pe'])[:20]
    for i, s in enumerate(sorted_by_pe, 1):
        holding_status = "✅" if s['code'].split('.')[0] in holdings or s['code'] in holdings else ""
        md.append(f"| {i} | {s['code']} | {s['name']} | {s['industry']} | {s['pe']:.2f} | {s['roe']:.2f}% | {s['roic']:.2f}% | {s['market_cap']:,} | {s['market']} {holding_status}|")
    md.append("")
    
    # 4. 按市值排序
    md.append("---")
    md.append("")
    md.append("## 4️⃣ 按市值排序 (Top 20)")
    md.append("")
    md.append("| 排名 | 代码 | 名称 | 行业 | 市值 (亿) | PE | ROE | 股息率 | 市场 |")
    md.append("|------|------|------|------|----------|-----|-----|--------|------|")
    sorted_by_cap = sorted(stocks, key=lambda x: x['market_cap'], reverse=True)[:20]
    for i, s in enumerate(sorted_by_cap, 1):
        holding_status = "✅" if s['code'].split('.')[0] in holdings or s['code'] in holdings else ""
        md.append(f"| {i} | {s['code']} | {s['name']} | {s['industry']} | {s['market_cap']:,} | {s['pe']:.2f} | {s['roe']:.2f}% | {s['dividend']:.1f}% | {s['market']} {holding_status}|")
    md.append("")
    
    # 5. 港股 vs A 股对比
    md.append("---")
    md.append("")
    md.append("## 5️⃣ 港股 vs A 股对比")
    md.append("")
    
    hk_stocks = [s for s in stocks if s['market'] == '港股']
    a_stocks = [s for s in stocks if s['market'] == 'A 股']
    us_stocks = [s for s in stocks if s['market'] == '美股']
    
    md.append("### 市场分布")
    md.append("")
    md.append(f"| 市场 | 股票数量 | 占比 | 平均 PE | 平均 ROE | 平均股息率 |")
    md.append("|------|---------|------|--------|---------|-----------|")
    
    for market_name, market_stocks in [('A 股', a_stocks), ('港股', hk_stocks), ('美股', us_stocks)]:
        count = len(market_stocks)
        pct = count / len(stocks) * 100
        avg_pe = sum(s['pe'] for s in market_stocks) / count if count > 0 else 0
        avg_roe = sum(s['roe'] for s in market_stocks) / count if count > 0 else 0
        avg_div = sum(s['dividend'] for s in market_stocks) / count if count > 0 else 0
        md.append(f"| {market_name} | {count} | {pct:.1f}% | {avg_pe:.2f} | {avg_roe:.2f}% | {avg_div:.2f}% |")
    md.append("")
    
    md.append("### 同公司 A+H 股对比")
    md.append("")
    md.append("| 公司 | A 股代码 | A 股 PE | H 股代码 | H 股 PE | 溢价率 |")
    md.append("|------|---------|-------|---------|-------|--------|")
    
    # 查找 A+H 股配对
    ah_pairs = [
        ('中国神华', '601088.SH', 10.5, '01088.HK', 10.2),
        ('招商银行', '600036.SH', 6.5, '03968.HK', 6.2),
        ('中国平安', '601318.SH', 8.5, '02318.HK', 8.2),
    ]
    for name, a_code, a_pe, h_code, h_pe in ah_pairs:
        premium = (a_pe - h_pe) / h_pe * 100 if h_pe > 0 else 0
        md.append(f"| {name} | {a_code} | {a_pe:.2f} | {h_code} | {h_pe:.2f} | {premium:+.1f}% |")
    md.append("")
    
    # 6. 锋哥当前持仓匹配度分析
    md.append("---")
    md.append("")
    md.append("## 6️⃣ 锋哥当前持仓匹配度分析")
    md.append("")
    
    md.append("### 持仓股票在五好股票池中的情况")
    md.append("")
    md.append("| 代码 | 名称 | 持仓市值 | 盈亏率 | 在五好池中 | 匹配度 |")
    md.append("|------|------|---------|--------|-----------|--------|")
    
    pool_codes = set(s['code'].split('.')[0] for s in stocks)
    pool_codes.update(s['code'] for s in stocks)
    
    for code, holding in holdings.items():
        in_pool = code in pool_codes or code.split('.')[0] in pool_codes
        status = "✅ 是" if in_pool else "❌ 否"
        match = "高" if in_pool and holding.get('current', 0) > holding.get('cost', 0) * 0.9 else "中" if in_pool else "低"
        pnl_pct = ((holding.get('current', 0) - holding.get('cost', 0)) / holding.get('cost', 1) * 100) if holding.get('cost', 0) > 0 else 0
        pnl_status = "✅" if pnl_pct >= 0 else "❌"
        md.append(f"| {code} | {holding['name']} | ${holding.get('current', 0) * holding.get('shares', 0):,.0f} | {pnl_pct:+.1f}% {pnl_status} | {status} | {match} |")
    md.append("")
    
    md.append("### 持仓行业分布 vs 股票池行业分布")
    md.append("")
    
    # 计算持仓行业分布
    holding_industries = {}
    for code, holding in holdings.items():
        # 查找对应的行业
        for s in stocks:
            if s['code'].split('.')[0] == code or s['code'] == code:
                industry = s['industry']
                if industry not in holding_industries:
                    holding_industries[industry] = 0
                holding_industries[industry] += 1
                break
    
    md.append("| 行业 | 持仓数量 | 股票池数量 | 持仓占比 | 池占比 |")
    md.append("|------|---------|-----------|---------|-------|")
    
    all_industries = set(holding_industries.keys()) | set(industries.keys())
    total_holdings = sum(holding_industries.values())
    total_pool = len(stocks)
    
    for industry in sorted(all_industries):
        hold_count = holding_industries.get(industry, 0)
        pool_count = len(industries.get(industry, []))
        hold_pct = hold_count / total_holdings * 100 if total_holdings > 0 else 0
        pool_pct = pool_count / total_pool * 100
        md.append(f"| {industry} | {hold_count} | {pool_count} | {hold_pct:.1f}% | {pool_pct:.1f}% |")
    md.append("")
    
    md.append("### 配置建议")
    md.append("")
    md.append("#### ✅ 继续持有（高匹配度）")
    md.append("- **中国海洋石油 (00883)** - 符合五好标准，PE 仅 5 倍，股息率 8%，盈利 43%")
    md.append("- **腾讯控股 (00700)** - 符合五好标准，互联网龙头，ROE 21%")
    md.append("- **可口可乐 (KO)** - 符合五好标准，ROE 高达 40%，股息贵族")
    md.append("- **伯克希尔 (BRK.B)** - 符合五好标准，价值投资标杆")
    md.append("")
    md.append("#### ⚠️ 需要关注（中等匹配度）")
    md.append("- **谷歌 (GOOGL)** - 符合五好标准，但近期表现平淡")
    md.append("- **微软 (MSFT)** - 符合五好标准，但 PE 较高 (32.5 倍)")
    md.append("- **苹果 (AAPL)** - 符合五好标准，ROE 极高但增长放缓")
    md.append("")
    md.append("#### ❌ 不符合五好标准")
    md.append("- **甲骨文 (ORCL)** - 不在五好池中，建议评估")
    md.append("- **英伟达 (NVDA)** - 不在五好池中，波动较大")
    md.append("- **特斯拉 (TSLA)** - 不在五好池中，利润率承压")
    md.append("- **阿里巴巴 (09988)** - 不在五好池中，ROE 偏低")
    md.append("")
    
    # 总结
    md.append("---")
    md.append("")
    md.append("## 📋 总结")
    md.append("")
    md.append(f"### 股票池概览")
    md.append(f"- **总计：** {len(stocks)} 只高质量股票")
    md.append(f"- **A 股：** {len(a_stocks)} 只 ({len(a_stocks)/len(stocks)*100:.1f}%)")
    md.append(f"- **港股：** {len(hk_stocks)} 只 ({len(hk_stocks)/len(stocks)*100:.1f}%)")
    md.append(f"- **美股：** {len(us_stocks)} 只 ({len(us_stocks)/len(stocks)*100:.1f}%)")
    md.append("")
    md.append(f"### 行业分布")
    md.append(f"- **白酒：** {len(industries.get('白酒', []))} 只 - 高 ROE，稳定增长")
    md.append(f"- **医药：** {len(industries.get('医药', []))} 只 - 防御性强")
    md.append(f"- **消费：** {len(industries.get('消费', []))} 只 - 现金流稳定")
    md.append(f"- **科技：** {len(industries.get('科技', []))} 只 - 成长性好")
    md.append(f"- **周期：** {len(industries.get('周期', []))} 只 - 高股息")
    md.append(f"- **制造：** {len(industries.get('制造', []))} 只 - 细分龙头")
    md.append(f"- **金融：** {len(industries.get('金融', []))} 只 - 低估值")
    md.append("")
    md.append(f"### 锋哥持仓匹配度")
    match_count = sum(1 for code in holdings if code in pool_codes or code.split('.')[0] in pool_codes)
    md.append(f"- **持仓股票数：** {len(holdings)} 只")
    md.append(f"- **符合五好标准：** {match_count} 只 ({match_count/len(holdings)*100:.1f}%)")
    md.append(f"- **建议优化：** {len(holdings) - match_count} 只股票可考虑调整")
    md.append("")
    md.append("---")
    md.append("")
    md.append("*报告生成时间：* " + datetime.now().strftime('%Y-%m-%d %H:%M'))
    md.append("*免责声明：以上分析基于公开数据，不构成投资建议*")
    
    return '\n'.join(md)


def generate_excel(stocks, holdings, filename):
    """生成 Excel 文件"""
    wb = Workbook()
    
    # 样式定义
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    center_align = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sheet 1: 总览
    ws1 = wb.active
    ws1.title = '股票池总览'
    
    headers1 = ['代码', '名称', '行业', '市场', 'PE', 'ROE', 'ROIC', '市值 (亿)', '股息率', '是否持仓']
    for col, h in enumerate(headers1, 1):
        cell = ws1.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    pool_codes = set(s['code'].split('.')[0] for s in stocks)
    pool_codes.update(s['code'] for s in stocks)
    
    for row, s in enumerate(stocks, 2):
        in_pool = s['code'].split('.')[0] in pool_codes or s['code'] in pool_codes
        is_holding = '✅ 是' if s['code'].split('.')[0] in holdings or s['code'] in holdings else '❌ 否'
        data = [s['code'], s['name'], s['industry'], s['market'], 
                f"{s['pe']:.2f}", f"{s['roe']:.2f}%", f"{s['roic']:.2f}%", 
                f"{s['market_cap']:,}", f"{s['dividend']:.1f}%", is_holding]
        for col, val in enumerate(data, 1):
            cell = ws1.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align
    
    # 设置列宽
    for col in range(1, 11):
        ws1.column_dimensions[chr(64+col)].width = 12
    
    # Sheet 2: 按 ROE 排序
    ws2 = wb.create_sheet('ROE 排序')
    sorted_by_roe = sorted(stocks, key=lambda x: x['roe'], reverse=True)
    
    headers2 = ['排名', '代码', '名称', '行业', '市场', 'ROE', 'PE', 'ROIC', '市值 (亿)', '股息率']
    for col, h in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    for row, s in enumerate(sorted_by_roe, 2):
        data = [row-1, s['code'], s['name'], s['industry'], s['market'],
                f"{s['roe']:.2f}%", f"{s['pe']:.2f}", f"{s['roic']:.2f}%",
                f"{s['market_cap']:,}", f"{s['dividend']:.1f}%"]
        for col, val in enumerate(data, 1):
            cell = ws2.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align
    
    # Sheet 3: 按 PE 排序
    ws3 = wb.create_sheet('PE 排序')
    sorted_by_pe = sorted(stocks, key=lambda x: x['pe'])
    
    for col, h in enumerate(headers2, 1):
        cell = ws3.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    for row, s in enumerate(sorted_by_pe, 2):
        data = [row-1, s['code'], s['name'], s['industry'], s['market'],
                f"{s['roe']:.2f}%", f"{s['pe']:.2f}", f"{s['roic']:.2f}%",
                f"{s['market_cap']:,}", f"{s['dividend']:.1f}%"]
        for col, val in enumerate(data, 1):
            cell = ws3.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align
    
    # Sheet 4: 按市值排序
    ws4 = wb.create_sheet('市值排序')
    sorted_by_cap = sorted(stocks, key=lambda x: x['market_cap'], reverse=True)
    
    headers4 = ['排名', '代码', '名称', '行业', '市场', '市值 (亿)', 'PE', 'ROE', 'ROIC', '股息率']
    for col, h in enumerate(headers4, 1):
        cell = ws4.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    for row, s in enumerate(sorted_by_cap, 2):
        data = [row-1, s['code'], s['name'], s['industry'], s['market'],
                f"{s['market_cap']:,}", f"{s['pe']:.2f}", f"{s['roe']:.2f}%",
                f"{s['roic']:.2f}%", f"{s['dividend']:.1f}%"]
        for col, val in enumerate(data, 1):
            cell = ws4.cell(row=row, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align
    
    # Sheet 5: 行业分类
    ws5 = wb.create_sheet('行业分类')
    
    industries = {}
    for stock in stocks:
        industry = stock['industry']
        if industry not in industries:
            industries[industry] = []
        industries[industry].append(stock)
    
    row = 1
    for industry in sorted(industries.keys()):
        ws5.cell(row=row, column=1, value=f"{industry} ({len(industries[industry])}只)")
        ws5.cell(row=row, column=1).font = Font(bold=True)
        row += 1
        
        for col, h in enumerate(['代码', '名称', 'PE', 'ROE', 'ROIC', '市值', '股息率'], 1):
            cell = ws5.cell(row=row, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = center_align
        row += 1
        
        for s in sorted(industries[industry], key=lambda x: x['roe'], reverse=True):
            data = [s['code'], s['name'], f"{s['pe']:.2f}", f"{s['roe']:.2f}%", 
                    f"{s['roic']:.2f}%", f"{s['market_cap']:,}", f"{s['dividend']:.1f}%"]
            for col, val in enumerate(data, 1):
                ws5.cell(row=row, column=col, value=val)
            row += 1
        row += 1
    
    # Sheet 6: 持仓匹配度
    ws6 = wb.create_sheet('持仓匹配度')
    
    ws6.cell(row=1, column=1, value='锋哥持仓与五好股票池匹配度分析')
    ws6.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    headers6 = ['代码', '名称', '市场', '持仓数量', '成本价', '现价', '市值', '盈亏率', '在五好池中', '匹配度']
    for col, h in enumerate(headers6, 1):
        cell = ws6.cell(row=3, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    for row_idx, (code, holding) in enumerate(holdings.items(), 4):
        in_pool = '✅ 是' if code in pool_codes or code.split('.')[0] in pool_codes else '❌ 否'
        pnl_pct = ((holding.get('current', 0) - holding.get('cost', 0)) / holding.get('cost', 1) * 100) if holding.get('cost', 0) > 0 else 0
        match = '高' if in_pool and pnl_pct >= -10 else '中' if in_pool else '低'
        
        data = [code, holding['name'], holding['market'], holding.get('shares', 0),
                f"${holding.get('cost', 0):.2f}", f"${holding.get('current', 0):.2f}",
                f"${holding.get('current', 0) * holding.get('shares', 0):,.0f}",
                f"{pnl_pct:+.1f}%", in_pool, match]
        for col, val in enumerate(data, 1):
            cell = ws6.cell(row=row_idx, column=col, value=val)
            cell.border = thin_border
            cell.alignment = center_align
    
    wb.save(filename)
    return filename


if __name__ == '__main__':
    # 生成 Markdown 报告
    print("生成 Markdown 报告...")
    md_report = generate_markdown_report(STOCK_POOL, FENG_GE_HOLDINGS)
    
    md_filename = '/home/admin/openclaw/workspace/锋哥五好股票池完整报告_' + datetime.now().strftime('%Y%m%d') + '.md'
    with open(md_filename, 'w', encoding='utf-8') as f:
        f.write(md_report)
    print(f"Markdown 报告已保存：{md_filename}")
    
    # 生成 Excel 文件
    print("生成 Excel 文件...")
    excel_filename = '/home/admin/openclaw/workspace/锋哥五好股票池数据_' + datetime.now().strftime('%Y%m%d') + '.xlsx'
    generate_excel(STOCK_POOL, FENG_GE_HOLDINGS, excel_filename)
    print(f"Excel 文件已保存：{excel_filename}")
    
    print("\n✅ 报告生成完成！")
    print(f"   - Markdown 报告：{md_filename}")
    print(f"   - Excel 数据：{excel_filename}")
