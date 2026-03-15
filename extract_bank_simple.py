#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
简单版：提取银行对账单 PDF 并转换为 Excel
"""

import subprocess
import openpyxl
import os
import re
from pathlib import Path

# PDF 文件目录
pdf_dir = Path("/home/admin/.openclaw/media/inbound")
# 输出 Excel 文件
output_excel = "/home/admin/openclaw/workspace/银行对账单汇总.xlsx"

# 获取所有对账单 PDF 文件
pdf_files = sorted([f for f in pdf_dir.glob("BOCVC442981877088*.pdf")])

print(f"找到 {len(pdf_files)} 个 PDF 文件")

# 创建 Excel 工作簿
wb = openpyxl.Workbook()
# 删除默认 sheet
wb.remove(wb.active)

# 汇总数据
all_transactions = []

for pdf_path in pdf_files:
    print(f"\n处理：{pdf_path.name}")
    
    # 从文件名提取年月信息
    parts = pdf_path.stem.split("-")
    if len(parts) >= 3:
        year = parts[1]
        seq = parts[2].lstrip("0")  # 去掉前导零
        sheet_name = f"{year}年{seq}月"
    else:
        sheet_name = pdf_path.stem[:15]
    
    # 创建新的工作表
    ws = wb.create_sheet(title=sheet_name)
    
    # 使用 pdftotext 提取文本
    try:
        result = subprocess.run(
            ["pdftotext", "-layout", str(pdf_path), "-"],
            capture_output=True,
            text=True,
            timeout=30
        )
        text = result.stdout
    except Exception as e:
        print(f"  错误：{e}")
        continue
    
    # 解析文本，提取表格行
    # 表格格式大致为：|序号 | 记账日 | 起息日 | 交易类型 | 凭证 | 用途/摘要 | 借方 | 贷方 | 余额 | 机构 | 备注 |
    lines = text.split("\n")
    
    row_num = 1
    headers_written = False
    
    for line in lines:
        # 跳过空行和分隔线
        if not line.strip() or line.strip().startswith("─"):
            continue
        
        # 检查是否是表头行
        if "序号" in line and "记账日" in line and "借方" in line:
            if not headers_written:
                headers = ["序号", "记账日", "起息日", "交易类型", "凭证", "用途/摘要", "借方发生额", "贷方发生额", "余额", "机构/柜员/流水", "备注"]
                for col_idx, header in enumerate(headers, 1):
                    ws.cell(row=row_num, column=col_idx, value=header)
                row_num += 1
                headers_written = True
            continue
        
        # 检查是否是数据行（以数字序号开头）
        match = re.match(r'^\s*(\d+)\s+(.+)', line)
        if match and headers_written:
            # 尝试按空格分割
            parts = line.split()
            if len(parts) >= 8:
                # 简化处理：将整行写入一个单元格，或者尝试分割
                # 由于格式复杂，我们写入主要字段
                ws.cell(row=row_num, column=1, value=row_num - 1)  # 序号
                ws.cell(row=row_num, column=2, value=line[:200])  # 整行内容放入摘要列
                
                # 尝试提取借贷方金额
                # 查找数字模式
                numbers = re.findall(r'[\d,]+\.?\d*', line)
                if len(numbers) >= 3:
                    # 假设最后三个数字是借方、贷方、余额
                    ws.cell(row=row_num, column=7, value=numbers[-3] if len(numbers) >= 3 else "")
                    ws.cell(row=row_num, column=8, value=numbers[-2] if len(numbers) >= 2 else "")
                    ws.cell(row=row_num, column=9, value=numbers[-1] if len(numbers) >= 1 else "")
                
                # 添加到汇总数据
                transaction = {
                    "文件": pdf_path.name,
                    "期间": sheet_name,
                    "原始数据": line[:500]
                }
                all_transactions.append(transaction)
                
                row_num += 1
    
    print(f"  写入 {row_num - 1} 行")

# 创建汇总工作表
summary_ws = wb.create_sheet(title="全部交易汇总", index=0)

# 写入汇总表头
summary_headers = ["文件", "期间", "原始数据"]
for col_idx, header in enumerate(summary_headers, 1):
    summary_ws.cell(row=1, column=col_idx, value=header)

# 写入汇总数据
for row_idx, trans in enumerate(all_transactions, 2):
    summary_ws.cell(row=row_idx, column=1, value=trans["文件"])
    summary_ws.cell(row=row_idx, column=2, value=trans["期间"])
    summary_ws.cell(row=row_idx, column=3, value=trans["原始数据"])

print(f"\n汇总交易记录：{len(all_transactions)} 条")

# 保存 Excel 文件
wb.save(output_excel)
print(f"\n✅ Excel 文件已保存：{output_excel}")
print(f"   文件大小：{os.path.getsize(output_excel) / 1024:.1f} KB")
