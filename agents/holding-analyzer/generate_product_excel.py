#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
生成私募基金产品持仓 Excel 报表
"""

import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
from pathlib import Path

# 产品数据（来自截图 OCR 识别 - 2026-03-26）
PRODUCTS = [
    # {产品全称，份额更新日期，最新份额，投资成本，最新市值，持仓收益}
    {
        'name': '前锋 1 号私募证券投资基金 A',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-03-26',
        'shares': 641672.01,
        'cost': 700000.00,
        'market_value': 720533.50,
        'profit': 20533.50,
        'return_rate': 2.93,
    },
    {
        'name': '领航 FOF1 号私募证券投资基金 A',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-01-12',
        'shares': 1300000.00,
        'cost': 1300000.00,
        'market_value': 1213940.00,
        'profit': -86060.00,
        'return_rate': -6.62,
    },
    {
        'name': '前沿 1 号私募证券投资基金',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2025-10-30',
        'shares': 2400000.00,
        'cost': 2400000.00,
        'market_value': 2499840.00,
        'profit': 99840.00,
        'return_rate': 4.16,
    },
    {
        'name': '前锋 3 号私募证券投资基金',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-03-26',
        'shares': 2304231.65,
        'cost': 2405830.87,
        'market_value': None,
        'profit': 101599.22,
        'return_rate': 4.41,
        'holding_days': 209,
    },
    {
        'name': '前锋 8 号私募证券投资基金',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-03-26',
        'shares': 800000.00,
        'cost': 797334.80,
        'market_value': None,
        'profit': -2665.19,
        'return_rate': -0.33,
        'holding_days': None,
    },
    {
        'name': '前锋 8 号私募证券投资基金',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-03-26',
        'shares': 481825.87,
        'cost': 500000.01,
        'market_value': None,
        'profit': 18174.14,
        'return_rate': 3.77,
        'holding_days': 275,
    },
    {
        'name': '前锋 8 号私募证券投资基金',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-03-26',
        'shares': 1018174.13,
        'cost': 1059556.06,
        'market_value': None,
        'profit': 41381.93,
        'return_rate': 4.06,
        'holding_days': 280,
    },
    {
        'name': '前锋 8 号私募证券投资基金',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-03-26',
        'shares': 100000.00,
        'cost': 104954.03,
        'market_value': None,
        'profit': 4954.03,
        'return_rate': 4.95,
        'holding_days': 228,
    },
    {
        'name': '前锋 8 号私募证券投资基金',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-03-26',
        'shares': 629788.35,
        'cost': 635489.91,
        'market_value': None,
        'profit': 5701.56,
        'return_rate': 0.91,
        'holding_days': 35,
    },
    {
        'name': '前锋 8 号私募证券投资基金',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-03-26',
        'shares': 1170211.65,
        'cost': 947788.45,
        'market_value': None,
        'profit': -222423.20,
        'return_rate': -19.01,
        'holding_days': 241,
    },
    {
        'name': '前锋 8 号私募证券投资基金',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-03-26',
        'shares': 500000.00,
        'cost': 402225.98,
        'market_value': None,
        'profit': -97774.02,
        'return_rate': -19.55,
        'holding_days': 197,
    },
    {
        'name': '乾享 1 号私募证券投资基金',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-03-26',
        'shares': 2000000.00,
        'cost': 1966955.15,
        'market_value': None,
        'profit': -33044.85,
        'return_rate': -1.65,
        'holding_days': None,
    },
    {
        'name': '乾享 1 号私募证券投资基金',
        'manager': '上海双城私募基金管理有限公司',
        'update_date': '2026-03-26',
        'shares': 1000000.00,
        'cost': 1001111.84,
        'market_value': None,
        'profit': 1111.84,
        'return_rate': 0.11,
        'holding_days': None,
    },
]

def create_product_excel(output_path):
    """创建产品持仓 Excel 报表"""
    
    # 创建工作簿
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "产品持仓明细"
    
    # 定义样式
    header_font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    cell_alignment = Alignment(horizontal='right', vertical='center')
    text_alignment = Alignment(horizontal='left', vertical='center')
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # 表头
    headers = [
        '序号', '产品名称', '管理人', '份额更新日期', '持有天数',
        '最新份额', '投资成本', '最新市值', '持仓收益', '收益率', '状态'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 数据行
    for row_idx, product in enumerate(PRODUCTS, 2):
        # 序号
        ws.cell(row=row_idx, column=1, value=row_idx - 1).alignment = Alignment(horizontal='center')
        
        # 产品名称
        ws.cell(row=row_idx, column=2, value=product['name']).alignment = text_alignment
        
        # 管理人
        ws.cell(row=row_idx, column=3, value=product['manager']).alignment = text_alignment
        
        # 份额更新日期
        ws.cell(row=row_idx, column=4, value=product['update_date']).alignment = Alignment(horizontal='center')
        
        # 持有天数
        holding_days = product.get('holding_days')
        ws.cell(row=row_idx, column=5, value=holding_days if holding_days else '-').alignment = Alignment(horizontal='center')
        
        # 最新份额
        ws.cell(row=row_idx, column=6, value=product['shares']).alignment = cell_alignment
        ws.cell(row=row_idx, column=6).number_format = '#,##0.00'
        
        # 投资成本
        ws.cell(row=row_idx, column=7, value=product['cost']).alignment = cell_alignment
        ws.cell(row=row_idx, column=7).number_format = '#,##0.00'
        
        # 最新市值
        market_value = product.get('market_value')
        if market_value:
            ws.cell(row=row_idx, column=8, value=market_value).alignment = cell_alignment
            ws.cell(row=row_idx, column=8).number_format = '#,##0.00'
        else:
            ws.cell(row=row_idx, column=8, value='-').alignment = Alignment(horizontal='center')
        
        # 持仓收益
        profit = product['profit']
        profit_cell = ws.cell(row=row_idx, column=9, value=profit)
        profit_cell.alignment = cell_alignment
        profit_cell.number_format = '#,##0.00'
        
        # 收益率
        return_rate = product.get('return_rate')
        return_cell = ws.cell(row=row_idx, column=10, value=return_rate)
        return_cell.alignment = Alignment(horizontal='right', vertical='center')
        return_cell.number_format = '0.00%'
        
        # 状态
        status = '盈利 ✅' if profit > 0 else ('亏损 ❌' if profit < 0 else '持平')
        status_cell = ws.cell(row=row_idx, column=11, value=status)
        status_cell.alignment = Alignment(horizontal='center')
        
        # 条件格式 - 盈亏颜色
        if profit > 0:
            profit_cell.font = Font(color='00B050')  # 绿色
            return_cell.font = Font(color='00B050')
        elif profit < 0:
            profit_cell.font = Font(color='FF0000')  # 红色
            return_cell.font = Font(color='FF0000')
        
        # 应用边框
        for col in range(1, 12):
            ws.cell(row=row_idx, column=col).border = thin_border
    
    # 汇总行
    total_row = len(PRODUCTS) + 2
    total_cost = sum(p['cost'] for p in PRODUCTS)
    total_profit = sum(p['profit'] for p in PRODUCTS)
    
    ws.cell(row=total_row, column=1, value='合计').alignment = Alignment(horizontal='right')
    ws.cell(row=total_row, column=7, value=total_cost).alignment = cell_alignment
    ws.cell(row=total_row, column=7).number_format = '#,##0.00'
    ws.cell(row=total_row, column=9, value=total_profit).alignment = cell_alignment
    ws.cell(row=total_row, column=9).number_format = '#,##0.00'
    ws.cell(row=total_row, column=9).font = Font(bold=True, color='00B050' if total_profit > 0 else 'FF0000')
    
    # 合并汇总行的前 6 列
    ws.merge_cells(f'A{total_row}:F{total_row}')
    
    # 列宽
    column_widths = [6, 35, 35, 14, 12, 18, 18, 18, 18, 12, 10]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # 冻结首行
    ws.freeze_panes = 'A2'
    
    # 保存文件
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    
    print(f"✅ Excel 报表已生成：{output_path}")
    print(f"   产品数量：{len(PRODUCTS)}")
    print(f"   总投资成本：¥{total_cost:,.2f}")
    print(f"   总持仓收益：¥{total_profit:,.2f}")
    
    return output_path

if __name__ == '__main__':
    output_file = Path(__file__).parent / 'data' / '产品持仓明细_20260326.xlsx'
    create_product_excel(output_file)
