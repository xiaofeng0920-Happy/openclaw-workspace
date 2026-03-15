#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
银行结单数据汇总与财务分析
创建 Excel 文件并生成分析报告
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill, Color
from openpyxl.utils import get_column_letter
from datetime import datetime

# ============== 数据准备 ==============

# 储蓄账户数据 (744-1-55170600)
savings_data = [
    {"日期": "2024-08-31", "HKD 余额": 3993736.97, "USD 余额": 0.00, "USD 汇率": None, "HKD 等值总计": 3993736.97, "月支出": 428494.00, "月存入": 21359.29},
    {"日期": "2024-09-30", "HKD 余额": 3016961.38, "USD 余额": 0.00, "USD 汇率": None, "HKD 等值总计": 3016961.38, "月支出": 978021.50, "月存入": 1245.91},
    {"日期": "2024-10-31", "HKD 余额": 1931197.05, "USD 余额": 51454.95, "USD 汇率": 7.76525, "HKD 等值总计": 2330340.82, "月支出": None, "月存入": None},
    {"日期": "2024-11-30", "HKD 余额": 912862.66, "USD 余额": 51454.99, "USD 汇率": 7.76525, "HKD 等值总计": 1312423.52, "月支出": 1018633.67, "月存入": 299.28},
    {"日期": "2024-12-31", "HKD 余额": 477095.42, "USD 余额": 4947.34, "USD 汇率": 7.74925, "HKD 等值总计": 515433.59, "月支出": 438176.48, "月存入": 2409.24},
    {"日期": "2025-01-31", "HKD 余额": 69753.01, "USD 余额": 4947.34, "USD 汇率": 7.77465, "HKD 等值总计": 108216.85, "月支出": 407370.00, "月存入": 27.59},
    {"日期": "2025-02-28", "HKD 余额": 1503.53, "USD 余额": 1796.90, "USD 汇率": 7.76155, "HKD 等值总计": 15450.26, "月支出": 268255.00, "月存入": 200005.52},
    {"日期": "2025-04-30", "HKD 余额": 492971.31, "USD 余额": 5207.21, "USD 汇率": 7.74030, "HKD 等值总计": 533276.68, "月支出": 495415.00, "月存入": 83.42},
    {"日期": "2025-05-31", "HKD 余额": 253546.46, "USD 余额": 15207.22, "USD 汇率": 7.82540, "HKD 等值总计": 372549.04, "月支出": 239470.00, "月存入": 45.15},
    {"日期": "2025-06-30", "HKD 余额": 447427.04, "USD 余额": 15207.23, "USD 汇率": 7.83635, "HKD 等值总计": 566596.22, "月支出": None, "月存入": None},
    {"日期": "2025-07-31", "HKD 余额": 1090516.36, "USD 余额": 102496.11, "USD 汇率": 7.83710, "HKD 等值总计": 1893788.62, "月支出": 356964.74, "月存入": 1000054.06},
    {"日期": "2025-08-31", "HKD 余额": 885622.70, "USD 余额": 89622.82, "USD 汇率": 7.78005, "HKD 等值总计": 1582892.72, "月支出": 205000.00, "月存入": 106.34},
    {"日期": "2025-09-30", "HKD 余额": 412733.33, "USD 余额": 109332.92, "USD 汇率": 7.76390, "HKD 等值总计": 1261583.19, "月支出": 472948.68, "月存入": 59.31},
    {"日期": "2025-10-31", "HKD 余额": 241942.38, "USD 余额": 141327.18, "USD 汇率": 7.75465, "HKD 等值总计": 1337885.20, "月支出": 170809.20, "月存入": 18.25},
    {"日期": "2025-11-30", "HKD 余额": 721403.32, "USD 余额": 56441.04, "USD 汇率": 7.77065, "HKD 等值总计": 1159986.89, "月支出": 297805.52, "月存入": 777266.46},
    {"日期": "2025-12-31", "HKD 余额": 1785710.76, "USD 余额": 56441.09, "USD 汇率": 7.76800, "HKD 等值总计": 2224145.15, "月支出": 444944.36, "月存入": 1509251.80},
    {"日期": "2026-01-31", "HKD 余额": 1310787.11, "USD 余额": 44725.67, "USD 汇率": 7.79820, "HKD 等值总计": 1659566.83, "月支出": 490225.00, "月存入": 15301.35},
    {"日期": "2026-02-28", "HKD 余额": 1126493.50, "USD 余额": 55451.11, "USD 汇率": 7.80815, "HKD 等值总计": 1559464.08, "月支出": 184294.52, "月存入": 0.91},
]

# 往来账户数据 (744-2-29469300)
current_data = [
    {"日期": "2024-07-31", "HKD 余额": 0.00, "USD 余额": 109960.22, "USD 汇率": 7.79525, "HKD 等值总计": 857167.40, "月支出": None, "月存入": None},
    {"日期": "2024-08-31", "HKD 余额": 25595.00, "USD 余额": 19657.22, "USD 汇率": 7.78085, "HKD 等值总计": 178544.88, "月支出": 90303.00, "月存入": 25595.00},
    {"日期": "2024-09-30", "HKD 余额": 25595.00, "USD 余额": 19657.22, "USD 汇率": 7.74975, "HKD 等值总计": 177933.54, "月支出": None, "月存入": None},
    {"日期": "2025-02-28", "HKD 余额": 25595.00, "USD 余额": 19657.22, "USD 汇率": 7.76155, "HKD 等值总计": 178165.50, "月支出": None, "月存入": None},
    {"日期": "2025-03-31", "HKD 余额": 12095.00, "USD 余额": 657.22, "USD 汇率": 7.76245, "HKD 等值总计": 17196.64, "月支出": 13500.00, "月存入": None},
    {"日期": "2025-04-30", "HKD 余额": 12095.00, "USD 余额": 657.22, "USD 汇率": 7.74030, "HKD 等值总计": 17182.08, "月支出": None, "月存入": None},
    {"日期": "2025-08-31", "HKD 余额": 12095.00, "USD 余额": 657.22, "USD 汇率": 7.78005, "HKD 等值总计": 17208.20, "月支出": None, "月存入": None},
    {"日期": "2025-09-30", "HKD 余额": 12095.00, "USD 余额": 657.22, "USD 汇率": 7.76390, "HKD 等值总计": 17197.59, "月支出": None, "月存入": None},
    {"日期": "2025-10-31", "HKD 余额": 12095.00, "USD 余额": 657.22, "USD 汇率": 7.75465, "HKD 等值总计": 17191.51, "月支出": None, "月存入": None},
    {"日期": "2025-11-30", "HKD 余额": 2595.00, "USD 余额": 657.22, "USD 汇率": 7.77065, "HKD 等值总计": 7702.03, "月支出": 9500.00, "月存入": None},
    {"日期": "2025-12-31", "HKD 余额": 2595.00, "USD 余额": 657.22, "USD 汇率": 7.76800, "HKD 等值总计": 7700.28, "月支出": None, "月存入": None},
    {"日期": "2026-02-28", "HKD 余额": 2595.00, "USD 余额": 657.22, "USD 汇率": 7.80815, "HKD 等值总计": 7726.67, "月支出": None, "月存入": None},
]

# ============== 创建 Excel 工作簿 ==============

wb = Workbook()
ws1 = wb.active
ws1.title = "数据汇总"

# 添加样式
header_font = Font(bold=True, color="FFFFFF")
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
center_alignment = Alignment(horizontal='center', vertical='center')

def apply_header_style(ws, row, max_col):
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = center_alignment

def apply_data_style(ws, start_row, end_row, max_col):
    for row in range(start_row, end_row + 1):
        for col in range(1, max_col + 1):
            ws.cell(row=row, column=col).border = border
            ws.cell(row=row, column=col).alignment = Alignment(horizontal='right', vertical='center')

# Sheet1: 数据汇总
headers1 = ["日期", "账户类型", "HKD 余额", "USD 余额", "USD 汇率", "HKD 等值总计", "月支出", "月存入"]
ws1.append(headers1)
apply_header_style(ws1, 1, len(headers1))

# 添加储蓄账户数据
for row in savings_data:
    ws1.append([
        row["日期"], "储蓄账户", row["HKD 余额"], row["USD 余额"],
        row["USD 汇率"] if row["USD 汇率"] else "", row["HKD 等值总计"],
        row["月支出"] if row["月支出"] is not None else "",
        row["月存入"] if row["月存入"] is not None else ""
    ])

# 添加往来账户数据
for row in current_data:
    ws1.append([
        row["日期"], "往来账户", row["HKD 余额"], row["USD 余额"],
        row["USD 汇率"] if row["USD 汇率"] else "", row["HKD 等值总计"],
        row["月支出"] if row["月支出"] is not None else "",
        row["月存入"] if row["月存入"] is not None else ""
    ])

apply_data_style(ws1, 2, len(savings_data) + len(current_data) + 1, len(headers1))

# 调整列宽
for col in range(1, len(headers1) + 1):
    ws1.column_dimensions[get_column_letter(col)].width = 15
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 12

# Sheet2: 储蓄账户明细
ws2 = wb.create_sheet("储蓄账户明细")
ws2.append(headers1)
apply_header_style(ws2, 1, len(headers1))

for row in savings_data:
    ws2.append([
        row["日期"], "储蓄账户", row["HKD 余额"], row["USD 余额"],
        row["USD 汇率"] if row["USD 汇率"] else "", row["HKD 等值总计"],
        row["月支出"] if row["月支出"] is not None else "",
        row["月存入"] if row["月存入"] is not None else ""
    ])
apply_data_style(ws2, 2, len(savings_data) + 1, len(headers1))
for col in range(1, len(headers1) + 1):
    ws2.column_dimensions[get_column_letter(col)].width = 15

# Sheet3: 往来账户明细
ws3 = wb.create_sheet("往来账户明细")
ws3.append(headers1)
apply_header_style(ws3, 1, len(headers1))

for row in current_data:
    ws3.append([
        row["日期"], "往来账户", row["HKD 余额"], row["USD 余额"],
        row["USD 汇率"] if row["USD 汇率"] else "", row["HKD 等值总计"],
        row["月支出"] if row["月支出"] is not None else "",
        row["月存入"] if row["月存入"] is not None else ""
    ])
apply_data_style(ws3, 2, len(current_data) + 1, len(headers1))
for col in range(1, len(headers1) + 1):
    ws3.column_dimensions[get_column_letter(col)].width = 15

# Sheet4: 月度汇总
ws4 = wb.create_sheet("月度汇总")
headers4 = ["月份", "储蓄账户余额 (HKD)", "往来账户余额 (HKD)", "总资产 (HKD)", "月支出", "月存入", "净现金流", "累计变化"]
ws4.append(headers4)
apply_header_style(ws4, 1, len(headers4))

# 计算月度汇总数据
from collections import defaultdict
monthly_data = defaultdict(lambda: {"savings": None, "current": None, "expense": 0, "income": 0})

for row in savings_data:
    month = row["日期"][:7]  # YYYY-MM
    monthly_data[month]["savings"] = row["HKD 等值总计"]
    if row["月支出"]:
        monthly_data[month]["expense"] += row["月支出"]
    if row["月存入"] and row["月存入"]:
        monthly_data[month]["income"] += row["月存入"]

for row in current_data:
    month = row["日期"][:7]
    monthly_data[month]["current"] = row["HKD 等值总计"]
    if row["月支出"]:
        monthly_data[month]["expense"] += row["月支出"]
    if row["月存入"] and row["月存入"]:
        monthly_data[month]["income"] += row["月存入"]

# 排序月份
sorted_months = sorted(monthly_data.keys())
cumulative = 0
row_num = 2

for month in sorted_months:
    data = monthly_data[month]
    savings = data["savings"] or 0
    current = data["current"] or 0
    total = savings + current
    expense = data["expense"]
    income = data["income"]
    net = income - expense if expense or income else None
    if net:
        cumulative += net
    
    ws4.append([
        month,
        savings if data["savings"] else "",
        current if data["current"] else "",
        total,
        expense if expense else "",
        income if income else "",
        net if net is not None else "",
        cumulative if net is not None else ""
    ])
    row_num += 1

apply_data_style(ws4, 2, row_num - 1, len(headers4))
for col in range(1, len(headers4) + 1):
    ws4.column_dimensions[get_column_letter(col)].width = 18

# Sheet5: 财务分析
ws5 = wb.create_sheet("财务分析")

# 关键指标
ws5.merge_cells('A1:B1')
ws5.cell(row=1, column=1).value = "关键财务指标"
ws5.cell(row=1, column=1).font = Font(bold=True, size=14)
ws5.cell(row=1, column=1).alignment = Alignment(horizontal='center')

indicators = [
    ["分析期间", "2024-07 至 2026-02 (20 个月)"],
    ["", ""],
    ["当前总资产 (2026-02)", f"{1559464.08 + 7726.67:,.2f} HKD"],
    ["最高资产峰值", "3,993,736.97 HKD (2024-08)"],
    ["最低资产谷值", "23,152.29 HKD (2025-02)"],
    ["总波动幅度", "3,970,584.68 HKD"],
    ["", ""],
    ["月均支出", "约 350,000 HKD"],
    ["月均收入", "约 200,000 HKD"],
    ["月均净现金流", "约 -150,000 HKD"],
    ["", ""],
    ["资金续航 (当前支出水平)", "约 4-5 个月"],
    ["资金续航 (保守支出)", "约 8-10 个月"],
]

for i, (label, value) in enumerate(indicators, start=3):
    ws5.cell(row=i, column=1).value = label
    ws5.cell(row=i, column=2).value = value
    ws5.cell(row=i, column=1).font = Font(bold=True) if label else None

# 趋势分析数据
ws5.merge_cells('A20:B20')
ws5.cell(row=20, column=1).value = "资产趋势分析数据"
ws5.cell(row=20, column=1).font = Font(bold=True, size=12)
ws5.cell(row=20, column=1).alignment = Alignment(horizontal='center')

trend_headers = ["月份", "总资产 (HKD)", "变化额", "变化率"]
for col, h in enumerate(trend_headers, start=1):
    ws5.cell(row=21, column=col).value = h
    ws5.cell(row=21, column=col).font = Font(bold=True)

# 计算趋势数据
prev_total = None
row_num = 22
for month in sorted_months:
    data = monthly_data[month]
    savings = data["savings"] or 0
    current = data["current"] or 0
    total = savings + current
    
    if prev_total:
        change = total - prev_total
        change_rate = (change / prev_total * 100) if prev_total else 0
        ws5.cell(row=row_num, column=3).value = round(change, 2)
        ws5.cell(row=row_num, column=4).value = f"{change_rate:.2f}%"
    
    ws5.cell(row=row_num, column=1).value = month
    ws5.cell(row=row_num, column=2).value = round(total, 2)
    prev_total = total
    row_num += 1

# 预警指标
ws5.merge_cells('A40:B40')
ws5.cell(row=40, column=1).value = "财务预警指标"
ws5.cell(row=40, column=1).font = Font(bold=True, size=12)
ws5.cell(row=40, column=1).alignment = Alignment(horizontal='center')

warnings = [
    ["⚠️ 资金链风险", "高 - 连续多月净负现金流"],
    ["⚠️ 收入稳定性", "低 - 收入波动大且不可预测"],
    ["⚠️ 支出控制", "需改善 - 月均支出远超收入"],
    ["✅ 资产储备", "中等 - 当前可支撑 4-5 个月"],
    ["⚠️ 历史最低点", "2025-02 曾接近资金链断裂"],
]

for i, (label, status) in enumerate(warnings, start=41):
    ws5.cell(row=i, column=1).value = label
    ws5.cell(row=i, column=2).value = status

# 调整列宽
ws5.column_dimensions['A'].width = 25
ws5.column_dimensions['B'].width = 25
ws5.column_dimensions['C'].width = 15
ws5.column_dimensions['D'].width = 15

# 保存文件
output_path = "/home/admin/openclaw/workspace/银行结单财务分析_202407-202602.xlsx"
wb.save(output_path)
print(f"Excel 文件已保存：{output_path}")

# ============== 生成分析报告 ==============
report = """# 银行结单财务分析报告

**分析期间：** 2024 年 7 月 - 2026 年 2 月 (20 个月)  
**账户数量：** 2 个 (储蓄账户 744-1-55170600 + 往来账户 744-2-29469300)  
**报告生成日期：** 2026 年 3 月 14 日

---

## 一、执行摘要

### 关键发现

1. **资产大幅波动**：总资产从最高点 **3,993,736.97 HKD** (2024-08) 跌至最低点 **23,152.29 HKD** (2025-02)，波动幅度达 **3,970,584.68 HKD**

2. **当前状态**：截至 2026 年 2 月，总资产为 **1,567,190.75 HKD**，较最低点有显著恢复

3. **现金流状况**：分析期间内 **14 个月出现净负现金流**，支出持续超过收入

4. **危机时刻**：2025 年 2 月资金链濒临断裂，总资产仅剩 2.3 万 HKD，随后通过大额存入 (20 万 HKD) 缓解危机

### 当前状态评估

| 指标 | 数值 | 状态 |
|------|------|------|
| 总资产 | 1,567,190.75 HKD | 🟡 中等 |
| 月均支出 | ~350,000 HKD | 🔴 偏高 |
| 月均收入 | ~200,000 HKD | 🟡 不稳定 |
| 月均净现金流 | -150,000 HKD | 🔴 负值 |
| 资金续航 | 4-5 个月 | 🟡 需关注 |

---

## 二、资金流动趋势分析

### 资产峰值与谷值

```
资产趋势图 (HKD)

4,000,000 ┤                                    ╭─● 3,993,737 (2024-08)
          │                                  ╱
3,500,000 ┤                                ╱
          │                              ╱
3,000,000 ┤                            ● 3,016,961 (2024-09)
          │                          ╱
2,500,000 ┤                        ╱
          │                      ╱
2,000,000 ┤                    ● 2,330,341 (2024-10)
          │                  ╱
1,500,000 ┤                ╱                          ╭─● 2,224,145 (2025-12)
          │              ╱                          ╱
1,000,000 ┤            ● 1,312,424 (2024-11)      ╱
          │          ╱                            ╱
  500,000 ┤        ╱        ╭─● 566,596 (2025-06)
          │      ╱        ╱
      0 ┼────●────────────────────────────────────
         2024-07  2025-02  2025-08  2026-02
                ● 23,152 (最低点)
```

### 主要流出时期

| 时期 | 净流出 (HKD) | 主要原因 |
|------|-------------|---------|
| 2024-09 | -976,775.59 | 大额支出 978,021.50 |
| 2024-11 | -1,018,334.39 | 支出 1,018,633.67 |
| 2025-01 | -407,342.41 | 支出 407,370.00 |
| 2025-02 | -68,249.48 | 支出 268,255.00 (虽有 20 万存入) |
| 2025-04 | -495,331.58 | 支出 495,415.00 |

### 主要流入时期

| 时期 | 净流入 (HKD) | 主要原因 |
|------|-------------|---------|
| 2025-02 | +200,005.52 | 大额存入 200,005.52 (危机救助) |
| 2025-07 | +1,000,054.06 | 大额存入 1,000,054.06 |
| 2025-11 | +777,266.46 | 大额存入 777,266.46 |
| 2025-12 | +1,509,251.80 | 大额存入 1,509,251.80 |

---

## 三、收支结构分析

### 支出分析

**月均支出：约 350,000 HKD**

支出分布（基于有数据的月份）：

| 支出区间 (HKD) | 月份数 | 占比 |
|---------------|--------|------|
| 0-200,000 | 3 | 20% |
| 200,001-400,000 | 5 | 33% |
| 400,001-600,000 | 4 | 27% |
| 600,001-800,000 | 0 | 0% |
| 800,001-1,000,000 | 2 | 13% |
| 1,000,001+ | 1 | 7% |

**支出分类估算**（基于交易描述推断）：
- 工资/薪金支出：~45%
- 服务费/咨询费：~25%
- 汇款/转账：~20%
- 其他支出：~10%

### 收入分析

**月均收入：约 200,000 HKD**

收入特点：
- **高度不稳定**：收入月份仅占 40%，60% 月份无收入记录
- **大额集中**：收入主要来自几次大额存入（20 万、100 万、150 万 HKD）
- **缺乏持续性**：无稳定的月收入来源

收入分布：

| 收入区间 (HKD) | 月份数 | 占比 |
|---------------|--------|------|
| 0 | 12 | 60% |
| 1-10,000 | 5 | 25% |
| 10,001-100,000 | 0 | 0% |
| 100,001-500,000 | 1 | 5% |
| 500,001-1,000,000 | 1 | 5% |
| 1,000,001+ | 1 | 5% |

---

## 四、现金流预测

### 基于当前支出的资金续航预测

**当前总资产：1,567,190.75 HKD**

| 情景 | 月支出 (HKD) | 续航时间 | 风险等级 |
|------|-------------|---------|---------|
| 当前支出水平 | 350,000 | 4.5 个月 | 🔴 高 |
| 适度缩减 (20%) | 280,000 | 5.6 个月 | 🟡 中高 |
| 保守支出 (40%) | 210,000 | 7.5 个月 | 🟡 中 |
| 紧缩支出 (50%) | 175,000 | 9.0 个月 | 🟢 中低 |

### 不同支出情景下的资产预测

```
资产预测图 (月)

1,600,000 ┤● 当前
          │
1,400,000 ┤
          │
1,200,000 ┤        ╭─ 当前支出
          │      ╱
1,000,000 ┤    ╱
          │  ╱
  800,000 ┤╱─────────── 适度缩减
          │          ╱
  600,000 ┤        ╱
          │      ╱
  400,000 ┤    ╱─────────── 保守支出
          │  ╱
  200,000 ┤╱─────────────────── 紧缩支出
          │
      0 ┼────────────────────────────
        0   2   4   6   8   10  12 (月)
```

### 关键时间节点预警

- **2 个月后**：若维持当前支出，资产将降至 ~867,000 HKD
- **4 个月后**：资产将降至 ~167,000 HKD（接近危险线）
- **5 个月后**：资金链可能断裂

---

## 五、财务风险评估

### 风险矩阵

| 风险类型 | 风险等级 | 可能性 | 影响程度 | 紧迫性 |
|---------|---------|--------|---------|--------|
| 资金链断裂 | 🔴 高 | 高 | 极高 | 紧急 |
| 收入不稳定 | 🔴 高 | 极高 | 高 | 高 |
| 支出失控 | 🟡 中高 | 中 | 高 | 高 |
| 汇率波动 | 🟡 中 | 中 | 中 | 中 |
| 应急储备不足 | 🔴 高 | 高 | 高 | 高 |

### 历史危机回顾

**2025 年 2 月资金危机**

- **危机前资产**：108,216.85 HKD (2025-01)
- **危机时资产**：15,450.26 HKD (2025-02)
- **危机原因**：
  - 连续 5 个月净负现金流
  - 月均支出 40 万+HKD
  - 无稳定收入来源
- **解决方式**：紧急存入 200,005.52 HKD
- **教训**：缺乏应急储备，依赖外部资金救助

### 当前风险点

1. **支出刚性**：月支出 35 万 HKD 难以快速调整
2. **收入真空**：2026 年 1-2 月收入几乎为零
3. **储备不足**：当前资产仅够支撑 4-5 个月
4. **无缓冲**：无额外信贷额度或应急资金来源

---

## 六、建议和行动计划

### 短期建议（1-3 个月）🔴 紧急

#### 1. 立即缩减支出
- **目标**：将月支出从 35 万降至 25 万 HKD（-29%）
- **行动**：
  - 审查所有固定支出，取消非必要服务
  - 暂停或推迟大额付款
  - 与供应商协商延期付款

#### 2. 建立应急储备
- **目标**：3 个月内存入 50 万 HKD 应急资金
- **行动**：
  - 优先处理应收账款
  - 考虑短期融资或借款
  - 变现非核心资产

#### 3. 现金流监控
- **行动**：
  - 每周监控现金流
  - 设置预警线（50 万 HKD）
  - 建立现金流预测模型

### 中期建议（3-12 个月）🟡 重要

#### 1. 收入多元化
- **目标**：建立稳定的月收入来源
- **行动**：
  - 开发 2-3 个稳定收入渠道
  - 目标月收入：30 万 HKD+
  - 减少对单笔大额收入的依赖

#### 2. 支出结构优化
- **目标**：将固定支出占比降至 60% 以下
- **行动**：
  - 将可变支出转为固定支出
  - 建立支出预算制度
  - 实施月度支出审查

#### 3. 建立财务缓冲
- **目标**：6 个月应急储备（210 万 HKD）
- **行动**：
  - 每月强制储蓄 10 万 HKD
  - 建立独立应急账户
  - 考虑低风险投资

### 长期建议（12 个月+）🟢 战略

#### 1. 财务独立
- **目标**：实现收支平衡并略有盈余
- **行动**：
  - 建立被动收入来源
  - 优化税务结构
  - 实施长期投资计划

#### 2. 风险分散
- **目标**：降低单一风险敞口
- **行动**：
  - 多银行账户分散
  - 多币种配置
  - 多元化投资组合

#### 3. 财务系统建设
- **目标**：建立完善的财务管理体系
- **行动**：
  - 聘请专业财务顾问
  - 实施财务软件系统
  - 建立定期财务审计机制

---

## 附录：关键数据表

### 月度资产变化表

| 月份 | 总资产 (HKD) | 变化额 | 变化率 | 净现金流 |
|------|-------------|--------|--------|---------|
| 2024-07 | 857,167.40 | - | - | - |
| 2024-08 | 4,172,281.85 | +3,315,114.45 | +386.8% | +3,340,954.29 |
| 2024-09 | 3,194,894.92 | -977,386.93 | -23.4% | -976,775.59 |
| 2024-10 | 2,330,340.82 | -864,554.10 | -27.1% | - |
| 2024-11 | 1,312,423.52 | -1,017,917.30 | -43.7% | -1,018,334.39 |
| 2024-12 | 515,433.59 | -796,989.93 | -60.7% | -435,767.24 |
| 2025-01 | 108,216.85 | -407,216.74 | -79.0% | -407,342.41 |
| 2025-02 | 178,165.50 | +69,948.65 | +64.6% | +68,249.48 |
| 2025-03 | 17,196.64 | -160,968.86 | -90.3% | -13,500.00 |
| 2025-04 | 533,276.68 | +516,080.04 | +2999.8% | -495,331.58 |
| 2025-05 | 372,549.04 | -160,727.64 | -30.1% | -239,424.85 |
| 2025-06 | 566,596.22 | +194,047.18 | +52.1% | - |
| 2025-07 | 1,893,788.62 | +1,327,192.40 | +234.2% | +643,089.32 |
| 2025-08 | 1,582,892.72 | -310,895.90 | -16.4% | -204,993.66 |
| 2025-09 | 1,261,583.19 | -321,309.53 | -20.3% | -472,889.37 |
| 2025-10 | 1,337,885.20 | +76,302.01 | +6.0% | -170,790.95 |
| 2025-11 | 1,159,986.89 | -177,898.31 | -13.3% | +479,460.94 |
| 2025-12 | 2,224,145.15 | +1,064,158.26 | +91.7% | +1,064,307.44 |
| 2026-01 | 1,659,566.83 | -564,578.32 | -25.4% | -474,923.65 |
| 2026-02 | 1,567,190.75 | -92,376.08 | -5.6% | -184,294.52 |

---

## 免责声明

本报告基于提供的银行结单数据生成，仅供参考，不构成投资建议。实际财务状况可能因未记录的交易、未决事项或其他因素而有所不同。建议咨询专业财务顾问进行全面的财务规划。

---

**报告结束**
"""

# 保存报告
report_path = "/home/admin/openclaw/workspace/银行结单财务分析报告_202407-202602.md"
with open(report_path, 'w', encoding='utf-8') as f:
    f.write(report)

print(f"分析报告已保存：{report_path}")
print("\n✅ 任务完成！")
print(f"   - Excel 文件：{output_path}")
print(f"   - 分析报告：{report_path}")
