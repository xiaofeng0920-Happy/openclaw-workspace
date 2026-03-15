#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
使用 camelot 提取银行对账单 PDF 表格并转换为 Excel
"""

import camelot
import openpyxl
import os
from pathlib import Path
from datetime import datetime

# PDF 文件目录
pdf_dir = Path("/home/admin/.openclaw/media/inbound")
# 输出 Excel 文件
output_excel = "/home/admin/openclaw/workspace/银行对账单汇总.xlsx"

# 获取所有对账单 PDF 文件
pdf_files = sorted([f for f in pdf_dir.glob("BOCVC442981877088*.pdf")])

print(f"找到 {len(pdf_files)} 个 PDF 文件")

# 创建 Excel 工作簿
wb = openpyxl.Workbook()
# 删除默认 sheet
wb.remove(wb.active)

# 汇总数据
all_transactions = []

for pdf_idx, pdf_path in enumerate(pdf_files):
    print(f"\n[{pdf_idx + 1}/{len(pdf_files)}] 处理：{pdf_path.name}")
    
    # 从文件名提取年月信息
    parts = pdf_path.stem.split("-")
    if len(parts) >= 3:
        year = parts[1]
        seq = parts[2].lstrip("0")
        sheet_name = f"{year}年{seq}月"
    else:
        sheet_name = pdf_path.stem[:15]
    
    # 限制工作表名称长度（Excel 限制 31 字符）
    sheet_name = sheet_name[:31]
    
    # 创建新的工作表
    ws = wb.create_sheet(title=sheet_name)
    
    try:
        # 使用 camelot 提取表格（lattice 模式适合有边框的表格）
        tables = camelot.read_pdf(str(pdf_path), pages='all', flavor='lattice')
        print(f"  提取到 {len(tables)} 个表格")
        
        row_num = 1
        
        for table_idx, table in enumerate(tables):
            df = table.df
            
            for row_idx, row in df.iterrows():
                values = row.tolist()
                
                # 检查是否是表头
                if any("序号" in str(v) for v in values) or any("No." in str(v) for v in values):
                    # 写入表头
                    for col_idx, value in enumerate(values, 1):
                        if pd.notna(value) and str(value).strip():
                            ws.cell(row=row_num, column=col_idx, value=str(value).strip())
                    row_num += 1
                    continue
                
                # 写入数据行
                for col_idx, value in enumerate(values, 1):
                    if pd.notna(value) and str(value).strip():
                        ws.cell(row=row_num, column=col_idx, value=str(value).strip())
                
                # 添加到汇总数据（如果是数据行）
                if len(values) >= 5:
                    transaction = {
                        "文件": pdf_path.name,
                        "期间": sheet_name,
                        "序号": values[0] if len(values) > 0 else "",
                        "记账日": values[1] if len(values) > 1 else "",
                        "起息日": values[2] if len(values) > 2 else "",
                        "交易类型": values[3] if len(values) > 3 else "",
                        "凭证": values[4] if len(values) > 4 else "",
                        "用途摘要": values[5] if len(values) > 5 else "",
                        "借方发生额": values[6] if len(values) > 6 else "",
                        "贷方发生额": values[7] if len(values) > 7 else "",
                        "余额": values[8] if len(values) > 8 else "",
                        "机构柜员": values[9] if len(values) > 9 else "",
                        "备注": values[10] if len(values) > 10 else "",
                    }
                    all_transactions.append(transaction)
                
                row_num += 1
        
        print(f"  写入 {row_num - 1} 行数据")
        
    except Exception as e:
        print(f"  错误：{e}")
        # 如果 lattice 模式失败，尝试 stream 模式
        try:
            tables = camelot.read_pdf(str(pdf_path), pages='all', flavor='stream')
            print(f"  (stream 模式) 提取到 {len(tables)} 个表格")
            
            row_num = 1
            for table_idx, table in enumerate(tables):
                df = table.df
                for row_idx, row in df.iterrows():
                    values = row.tolist()
                    for col_idx, value in enumerate(values, 1):
                        if pd.notna(value) and str(value).strip():
                            ws.cell(row=row_num, column=col_idx, value=str(value).strip())
                    row_num += 1
            print(f"  写入 {row_num - 1} 行数据")
        except Exception as e2:
            print(f"  stream 模式也失败：{e2}")

# 创建汇总工作表
summary_ws = wb.create_sheet(title="全部交易汇总", index=0)

# 写入汇总表头
summary_headers = ["文件", "期间", "序号", "记账日", "起息日", "交易类型", "凭证", "用途摘要", 
                   "借方发生额", "贷方发生额", "余额", "机构柜员", "备注"]
for col_idx, header in enumerate(summary_headers, 1):
    summary_ws.cell(row=1, column=col_idx, value=header)

# 写入汇总数据
for row_idx, trans in enumerate(all_transactions, 2):
    summary_ws.cell(row=row_idx, column=1, value=trans["文件"])
    summary_ws.cell(row=row_idx, column=2, value=trans["期间"])
    summary_ws.cell(row=row_idx, column=3, value=trans["序号"])
    summary_ws.cell(row=row_idx, column=4, value=trans["记账日"])
    summary_ws.cell(row=row_idx, column=5, value=trans["起息日"])
    summary_ws.cell(row=row_idx, column=6, value=trans["交易类型"])
    summary_ws.cell(row=row_idx, column=7, value=trans["凭证"])
    summary_ws.cell(row=row_idx, column=8, value=trans["用途摘要"])
    summary_ws.cell(row=row_idx, column=9, value=trans["借方发生额"])
    summary_ws.cell(row=row_idx, column=10, value=trans["贷方发生额"])
    summary_ws.cell(row=row_idx, column=11, value=trans["余额"])
    summary_ws.cell(row=row_idx, column=12, value=trans["机构柜员"])
    summary_ws.cell(row=row_idx, column=13, value=trans["备注"])

print(f"\n汇总交易记录：{len(all_transactions)} 条")

# 保存 Excel 文件
wb.save(output_excel)
print(f"\n✅ Excel 文件已保存：{output_excel}")
print(f"   文件大小：{os.path.getsize(output_excel) / 1024:.1f} KB")
