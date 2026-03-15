#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
银行对账单 PDF 转 Excel - 文本解析版
"""

import subprocess
import openpyxl
import re
import os
from pathlib import Path
from datetime import datetime

pdf_dir = Path("/home/admin/.openclaw/media/inbound")
output_excel = "/home/admin/openclaw/workspace/银行对账单汇总.xlsx"

pdf_files = sorted(pdf_dir.glob("BOCVC442981877088*.pdf"))
print(f"找到 {len(pdf_files)} 个 PDF 文件")

wb = openpyxl.Workbook()
wb.remove(wb.active)

all_transactions = []

for pdf_idx, pdf_path in enumerate(pdf_files):
    print(f"\n[{pdf_idx + 1}/{len(pdf_files)}] {pdf_path.name}")
    
    # 从文件名提取期间
    parts = pdf_path.stem.split("-")
    if len(parts) >= 3:
        year = parts[1]
        seq = parts[2].lstrip("0")
        sheet_name = f"{year}年{seq}月"
    else:
        sheet_name = "unknown"
    sheet_name = sheet_name[:31]
    
    ws = wb.create_sheet(sheet_name)
    
    # 写入表头
    headers = ["序号", "记账日", "起息日", "交易类型", "凭证", "用途/摘要", 
               "借方发生额", "贷方发生额", "余额", "机构/柜员", "备注"]
    for ci, h in enumerate(headers, 1):
        ws.cell(1, ci, h)
    
    # 提取 PDF 文本
    try:
        result = subprocess.run(
            ["pdftotext", "-layout", str(pdf_path), "-"],
            capture_output=True, text=True, timeout=30
        )
        text = result.stdout
    except Exception as e:
        print(f"  错误：{e}")
        continue
    
    row_num = 2
    lines = text.split("\n")
    
    for line in lines:
        # 跳过空行、分隔线、表头、汇总行
        if not line.strip():
            continue
        if "─" in line or "═" in line:
            continue
        if "序号" in line and "记账日" in line:
            continue
        if "借方合计" in line or "贷方合计" in line:
            continue
        if "Account No" in line or "币种" in line:
            continue
        if "页" in line and "共" in line:
            continue
        
        # 尝试匹配数据行（以数字序号开头，如 | 1   |250207|）
        # 格式：| 序号 | 记账日 | 起息日 | 交易类型 | 凭证 | 摘要 | 借方 | 贷方 | 余额 | 机构 | 备注 |
        match = re.search(r'\|\s*(\d+)\s*\|', line)
        if match:
            seq_num = match.group(1)
            
            # 按 | 分割
            parts = line.split("|")
            if len(parts) >= 9:
                # 清理数据
                cleaned = [p.strip() for p in parts]
                
                # 提取字段
                data = {
                    "序号": cleaned[1] if len(cleaned) > 1 else "",
                    "记账日": cleaned[2] if len(cleaned) > 2 else "",
                    "起息日": cleaned[3] if len(cleaned) > 3 else "",
                    "交易类型": cleaned[4] if len(cleaned) > 4 else "",
                    "凭证": cleaned[5] if len(cleaned) > 5 else "",
                    "用途摘要": cleaned[6] if len(cleaned) > 6 else "",
                    "借方": cleaned[7] if len(cleaned) > 7 else "",
                    "贷方": cleaned[8] if len(cleaned) > 8 else "",
                    "余额": cleaned[9] if len(cleaned) > 9 else "",
                    "机构": cleaned[10] if len(cleaned) > 10 else "",
                    "备注": cleaned[11] if len(cleaned) > 11 else "",
                }
                
                # 写入 Excel
                ws.cell(row_num, 1, data["序号"])
                ws.cell(row_num, 2, data["记账日"])
                ws.cell(row_num, 3, data["起息日"])
                ws.cell(row_num, 4, data["交易类型"])
                ws.cell(row_num, 5, data["凭证"])
                ws.cell(row_num, 6, data["用途摘要"])
                ws.cell(row_num, 7, data["借方"])
                ws.cell(row_num, 8, data["贷方"])
                ws.cell(row_num, 9, data["余额"])
                ws.cell(row_num, 10, data["机构"])
                ws.cell(row_num, 11, data["备注"])
                
                # 添加到汇总
                all_transactions.append({
                    "文件": pdf_path.name,
                    "期间": sheet_name,
                    **data
                })
                
                row_num += 1
    
    print(f"  提取 {row_num - 2} 条记录")

# 创建汇总表
summary_ws = wb.create_sheet("全部交易汇总", index=0)
sum_headers = ["文件", "期间", "序号", "记账日", "起息日", "交易类型", "凭证", 
               "用途摘要", "借方发生额", "贷方发生额", "余额", "机构柜员", "备注"]
for ci, h in enumerate(sum_headers, 1):
    summary_ws.cell(1, ci, h)

for ri, t in enumerate(all_transactions, 2):
    summary_ws.cell(ri, 1, t["文件"])
    summary_ws.cell(ri, 2, t["期间"])
    summary_ws.cell(ri, 3, t["序号"])
    summary_ws.cell(ri, 4, t["记账日"])
    summary_ws.cell(ri, 5, t["起息日"])
    summary_ws.cell(ri, 6, t["交易类型"])
    summary_ws.cell(ri, 7, t["凭证"])
    summary_ws.cell(ri, 8, t["用途摘要"])
    summary_ws.cell(ri, 9, t["借方"])
    summary_ws.cell(ri, 10, t["贷方"])
    summary_ws.cell(ri, 11, t["余额"])
    summary_ws.cell(ri, 12, t["机构"])
    summary_ws.cell(ri, 13, t["备注"])

wb.save(output_excel)
print(f"\n✅ 完成！")
print(f"   总记录数：{len(all_transactions)} 条")
print(f"   文件：{output_excel}")
print(f"   大小：{os.path.getsize(output_excel)/1024:.1f} KB")
