#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
提取银行对账单 PDF 并转换为 Excel
"""

import pdfplumber
import openpyxl
import os
from pathlib import Path
from datetime import datetime

# PDF 文件目录
pdf_dir = Path("/home/admin/.openclaw/media/inbound")
# 输出 Excel 文件
output_excel = "/home/admin/openclaw/workspace/银行对账单汇总.xlsx"

# 获取所有对账单 PDF 文件
pdf_files = sorted([f for f in pdf_dir.glob("BOCVC442981877088*.pdf")])

print(f"找到 {len(pdf_files)} 个 PDF 文件")

# 创建 Excel 工作簿
wb = openpyxl.Workbook()
# 删除默认 sheet
wb.remove(wb.active)

# 汇总数据
all_transactions = []

for pdf_path in pdf_files:
    print(f"\n处理：{pdf_path.name}")
    
    # 从文件名提取年月信息
    # 格式：BOCVC442981877088-YYYY-NNNN.pdf
    parts = pdf_path.stem.split("-")
    if len(parts) >= 3:
        year = parts[1]
        seq = parts[2]
        sheet_name = f"{year}年{seq}月"
    else:
        sheet_name = pdf_path.stem[:20]
    
    # 创建新的工作表
    ws = wb.create_sheet(title=sheet_name)
    
    # 读取 PDF
    with pdfplumber.open(pdf_path) as pdf:
        print(f"  共 {len(pdf.pages)} 页")
        
        row_num = 1
        col_headers = None
        
        for page_num, page in enumerate(pdf.pages):
            print(f"  处理第 {page_num + 1} 页...")
            
            # 提取表格
            tables = page.extract_tables()
            
            for table in tables:
                for row_idx, row in enumerate(table):
                    if row is None:
                        continue
                    
                    # 清理数据
                    cleaned_row = []
                    for cell in row:
                        if cell is None:
                            cleaned_row.append("")
                        else:
                            # 清理文本，替换换行符
                            cell_text = str(cell).replace("\n", " ").strip()
                            cleaned_row.append(cell_text)
                    
                    # 检查是否是表头
                    if col_headers is None and ("序号" in str(cleaned_row) or "No." in str(cleaned_row)):
                        col_headers = cleaned_row
                        # 写入表头
                        for col_idx, header in enumerate(cleaned_row, 1):
                            ws.cell(row=row_num, column=col_idx, value=header)
                        row_num += 1
                    elif col_headers is not None:
                        # 写入数据行
                        for col_idx, value in enumerate(cleaned_row, 1):
                            if value:  # 只写入非空值
                                ws.cell(row=row_num, column=col_idx, value=value)
                        
                        # 同时添加到汇总数据
                        if len(cleaned_row) >= 10:  # 确保是数据行
                            transaction = {
                                "文件": pdf_path.name,
                                "期间": sheet_name,
                                "序号": cleaned_row[0] if len(cleaned_row) > 0 else "",
                                "记账日": cleaned_row[1] if len(cleaned_row) > 1 else "",
                                "起息日": cleaned_row[2] if len(cleaned_row) > 2 else "",
                                "交易类型": cleaned_row[3] if len(cleaned_row) > 3 else "",
                                "凭证": cleaned_row[4] if len(cleaned_row) > 4 else "",
                                "用途摘要": cleaned_row[5] if len(cleaned_row) > 5 else "",
                                "借方发生额": cleaned_row[6] if len(cleaned_row) > 6 else "",
                                "贷方发生额": cleaned_row[7] if len(cleaned_row) > 7 else "",
                                "余额": cleaned_row[8] if len(cleaned_row) > 8 else "",
                                "机构柜员": cleaned_row[9] if len(cleaned_row) > 9 else "",
                                "备注": cleaned_row[10] if len(cleaned_row) > 10 else "",
                            }
                            all_transactions.append(transaction)
                        
                        row_num += 1
        
        print(f"  写入 {row_num - 1} 行")

# 创建汇总工作表
summary_ws = wb.create_sheet(title="全部交易汇总", index=0)

# 写入汇总表头
summary_headers = ["文件", "期间", "序号", "记账日", "起息日", "交易类型", "凭证", "用途摘要", 
                   "借方发生额", "贷方发生额", "余额", "机构柜员", "备注"]
for col_idx, header in enumerate(summary_headers, 1):
    summary_ws.cell(row=1, column=col_idx, value=header)

# 写入汇总数据
for row_idx, trans in enumerate(all_transactions, 2):
    summary_ws.cell(row=row_idx, column=1, value=trans["文件"])
    summary_ws.cell(row=row_idx, column=2, value=trans["期间"])
    summary_ws.cell(row=row_idx, column=3, value=trans["序号"])
    summary_ws.cell(row=row_idx, column=4, value=trans["记账日"])
    summary_ws.cell(row=row_idx, column=5, value=trans["起息日"])
    summary_ws.cell(row=row_idx, column=6, value=trans["交易类型"])
    summary_ws.cell(row=row_idx, column=7, value=trans["凭证"])
    summary_ws.cell(row=row_idx, column=8, value=trans["用途摘要"])
    summary_ws.cell(row=row_idx, column=9, value=trans["借方发生额"])
    summary_ws.cell(row=row_idx, column=10, value=trans["贷方发生额"])
    summary_ws.cell(row=row_idx, column=11, value=trans["余额"])
    summary_ws.cell(row=row_idx, column=12, value=trans["机构柜员"])
    summary_ws.cell(row=row_idx, column=13, value=trans["备注"])

print(f"\n汇总交易记录：{len(all_transactions)} 条")

# 保存 Excel 文件
wb.save(output_excel)
print(f"\n✅ Excel 文件已保存：{output_excel}")

# 生成统计信息
print("\n" + "="*60)
print("📊 统计信息")
print("="*60)

# 按期间统计
period_stats = {}
for trans in all_transactions:
    period = trans["期间"]
    if period not in period_stats:
        period_stats[period] = {"debit": 0, "credit": 0, "count": 0}
    period_stats[period]["count"] += 1
    
    # 尝试解析金额
    try:
        debit_str = trans["借方发生额"].replace(",", "").strip() if trans["借方发生额"] else "0"
        if debit_str:
            period_stats[period]["debit"] += float(debit_str)
    except:
        pass
    
    try:
        credit_str = trans["贷方发生额"].replace(",", "").strip() if trans["贷方发生额"] else "0"
        if credit_str:
            period_stats[period]["credit"] += float(credit_str)
    except:
        pass

print(f"\n共有 {len(period_stats)} 个期间的数据")
print("\n期间统计:")
for period in sorted(period_stats.keys()):
    stats = period_stats[period]
    print(f"  {period}: {stats['count']} 条记录，借方 {stats['debit']:,.2f} 元，贷方 {stats['credit']:,.2f} 元")

print("\n✅ 转换完成！")
